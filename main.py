from pathlib import Path
import os
import sys
import subprocess
import json
from datetime import date, datetime
import math
import re
import calendar as _calendar
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.legend import Legend  # added legend import
    try:
        # Data labels for better readability (percentages inside slices)
        from openpyxl.chart.label import DataLabelList
    except Exception:
        DataLabelList = None
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel functionality disabled.")
from tkinter import (
    Button,
    Canvas,
    Entry,
    Frame,
    Label,
    Listbox,
    messagebox,
    PhotoImage,
    StringVar,
    Tk,
    TclError,
    Toplevel,
    filedialog,
)
from tkinter import ttk
from tkinter import font as tkfont

ICON_SUBSAMPLE = 2
ENTRY_NAMES = ["Date", "Libellé", "Montant", "Catégorie"]
CATEGORY_VALUES = [
    "Loyer","Courses","Loisirs","Transport","Santé","Abonnements","Restaurants","Cadeaux","Épargne","Salaire","Remboursement","Autre",
]
FONT_MIN_SIZE = 10
FONT_MAX_SIZE = 18
FONT_SCALE_DIVISOR = 35
FONT_FAMILY = "Segoe UI"

# Excel export styling constants (adjusted per user request)
EXCEL_BODY_FONT_SIZE = 16
EXCEL_TITLE_FONT_SIZE = 18

# Chart sizing (reduced by factor ~1.5 from previous 24x18)
CHART_WIDTH = 16  # was 24
CHART_HEIGHT = 12  # was 18

# Dynamic sizing constants for toggle switches
TOGGLE_MIN_HEIGHT = 20
TOGGLE_MAX_HEIGHT = 56  # Prevents oversized toggles on huge windows
TOGGLE_ASPECT_RATIO = 1.9  # width = height * ratio

# Path display left margin tuning (pixels)
PATH_LEFT_MARGIN_WIDE = 80
PATH_LEFT_MARGIN_NARROW = 50

LIGHT_THEME = {
    "bg": "#f0f0f0",
    "fg": "#333333",
    "entry_bg": "#ffffff",
    "entry_fg": "#333333",
    "entry_border": "#cccccc",
    "toggle_track": "#d32f2f",
    "toggle_track_active": "#4caf50",
    "toggle_thumb": "#ffffff",
}

DARK_THEME = {
    "bg": "#2e2e2e",
    "fg": "#f0f0f0",
    "entry_bg": "#3a3a3a",
    "entry_fg": "#f0f0f0",
    "entry_border": "#5a5a5a",
    "toggle_track": "#b71c1c",
    "toggle_track_active": "#66bb6a",
    "toggle_thumb": "#1e0303",
}


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Compta App")

        width_size = self.root.winfo_screenwidth()
        height_size = self.root.winfo_screenheight()
        # Initial window size set to half of the screen dimensions
        win_w = width_size // 2
        win_h = height_size // 2
        # Compute centered position (accounting for full screen size)
        pos_x = (width_size - win_w) // 2
        pos_y = (height_size - win_h) // 2
        self.root.geometry(f"{win_w}x{win_h}+{pos_x}+{pos_y}")
        self.root.minsize(400, 300)

        # Helper to re-center window if needed later (not auto-called to avoid disrupting user)
        def _center_root_once():
            try:
                self.root.update_idletasks()
                cur_w = self.root.winfo_width()
                cur_h = self.root.winfo_height()
                sw = self.root.winfo_screenwidth()
                sh = self.root.winfo_screenheight()
                cx = (sw - cur_w) // 2
                cy = (sh - cur_h) // 2
                self.root.geometry(f"{cur_w}x{cur_h}+{cx}+{cy}")
            except Exception:
                pass
        # Store for potential future use
        self._center_root_once = _center_root_once

        asset_dir = Path(__file__).resolve().parent / "src"
        dark_icon = PhotoImage(master=self.root, file=str(asset_dir / "dark_mode_icon.png"))
        light_icon = PhotoImage(master=self.root, file=str(asset_dir / "light_mode_icon.png"))
        try:
            self.app_icon = PhotoImage(master=self.root, file=str(asset_dir / "compta_app_logo.png"))
            self.root.iconphoto(True, self.app_icon)
        except Exception:
            self.app_icon = None
        self.dark_mode_icon = dark_icon.subsample(ICON_SUBSAMPLE, ICON_SUBSAMPLE)
        self.light_mode_icon = light_icon.subsample(ICON_SUBSAMPLE, ICON_SUBSAMPLE)

        self.is_dark_mode = False
        import json
        self.current_theme = LIGHT_THEME

        self.label_font = tkfont.Font(family=FONT_FAMILY, size=FONT_MIN_SIZE)
        self.entry_font = tkfont.Font(family=FONT_FAMILY, size=FONT_MIN_SIZE)
        # Slightly smaller font dedicated to combobox dropdown list items to keep the dropdown compact
        try:
            # Start a bit smaller than entry font; will be kept ~2pt below dynamically
            self.dropdown_font = tkfont.Font(family=FONT_FAMILY, size=max(FONT_MIN_SIZE - 2, 8))
        except Exception:
            # Fallback to entry font if creation fails
            self.dropdown_font = self.entry_font

        self.style = ttk.Style(self.root)
        try:
            self.style.theme_use("clam")
        except TclError:
            pass
        self.combobox_style = "App.TCombobox"
        # Compact combobox style for header account selector
        self.header_combobox_style = "Header.TCombobox"
        # Font for compact header combobox (initialized below in _update_fonts)
        self.header_cb_font = tkfont.Font(family=FONT_FAMILY, size=FONT_MIN_SIZE)
        # Smaller font for path action buttons (Ouvrir / Changer)
        self.small_button_font = tkfont.Font(family=FONT_FAMILY, size=max(FONT_MIN_SIZE - 4, 7))

        self.labels = []
        self.entries = []
        self.category_var = StringVar()
        self.category_combobox = None
        self.libelle_var = StringVar()
        self.montant_var = StringVar()
        self.amount_validator = self.root.register(self._validate_amount)
        self.libelle_placeholder = "Nom de la transaction"
        self.libelle_has_placeholder = True
        self.montant_placeholder = "0.00"
        self.montant_has_placeholder = True
        self.transaction_is_entry = False
        self.transaction_container = None
        self.is_prelevement = False
        self.prelevement_container = None
        self.date_var = StringVar(value=date.today().strftime("%d-%m-%Y"))
        self.date_entry = None
        self.calendar_win = None
        self._calendar_year = date.today().year
        self._calendar_month = date.today().month
        self._selected_date = None  # Track the selected date for highlighting
        self._calendar_opening = False  # Flag to prevent immediate closing

        self.header_frame = Frame(self.root, bg=LIGHT_THEME["bg"])
        self.header_frame.pack(fill="x")

        # Burger menu button (replaces theme toggle)
        self.menu_button = Button(
            self.header_frame,
            text="☰",
            command=self._open_menu,
            highlightthickness=0,
            borderwidth=0,
            relief="flat",
            bg=LIGHT_THEME["bg"],
            fg=LIGHT_THEME["fg"],
            activebackground=LIGHT_THEME["entry_bg"],
            activeforeground=LIGHT_THEME["fg"],
            cursor="hand2",
            font=(FONT_FAMILY, 16, "bold"),
        )
        self.menu_button.pack(side="right", padx=12, pady=8)

        # Prepare (hidden) overlay menu; built lazily on first open
        self.menu_overlay = None
        self.menu_close_button = None
        self.menu_center_frame = None
        self.menu_title_label = None
        self.menu_theme_button = None

        # (Removed 'Comptes...' button per user request)

        self.body = Frame(self.root, bg=LIGHT_THEME["bg"])
        self.body.pack(expand=True, fill="both", padx=40, pady=10)

        self.body.grid_rowconfigure(0, weight=1)
        self.body.grid_rowconfigure(1, weight=0)
        self.body.grid_rowconfigure(2, weight=1)
        self.body.grid_columnconfigure(0, weight=1)

        self.entry_frame = Frame(self.body, bg=LIGHT_THEME["bg"])
        self.entry_frame.grid(row=1, column=0)

        for i in range(4):
            self.entry_frame.columnconfigure(i, weight=1, uniform="col")

        for col, name in enumerate(ENTRY_NAMES):
            lbl = Label(
                self.entry_frame,
                text=name,
                anchor="center",
                pady=0,
                bg=LIGHT_THEME["bg"],
                fg=LIGHT_THEME["fg"],
                font=self.label_font,
            )
            lbl.grid(row=0, column=col, sticky="ew", padx=10, pady=(0, 4))
            self.labels.append(lbl)

        self.date_entry = Entry(
            self.entry_frame,
            textvariable=self.date_var,
            justify="center",
            relief="flat",
            highlightthickness=1,
            font=self.entry_font,
            exportselection=False,
        )
        self.date_entry.grid(row=1, column=0, sticky="new", padx=10, pady=(0, 8), ipady=4)
        self.date_entry.bind("<Button-1>", lambda e: self._on_date_entry_click())
        self.entries.append(self.date_entry)

        libelle_entry = Entry(
            self.entry_frame,
            textvariable=self.libelle_var,
            justify="center",
            relief="flat",
            highlightthickness=1,
            font=self.entry_font,
            exportselection=False,
        )
        libelle_entry.grid(row=1, column=1, sticky="new", padx=10, pady=(0, 8), ipady=4)
        libelle_entry.bind("<FocusIn>", self._on_libelle_focus_in)
        libelle_entry.bind("<FocusOut>", self._on_libelle_focus_out)
        self.entries.append(libelle_entry)
        self.libelle_entry = libelle_entry

        montant_entry = Entry(
            self.entry_frame,
            textvariable=self.montant_var,
            justify="center",
            relief="flat",
            highlightthickness=1,
            font=self.entry_font,
            exportselection=False,
            validate="key",
            validatecommand=(self.amount_validator, "%P"),
        )
        montant_entry.grid(row=1, column=2, sticky="new", padx=10, pady=(0, 8), ipady=4)
        montant_entry.bind("<FocusIn>", self._on_montant_focus_in)
        montant_entry.bind("<FocusOut>", self._on_montant_focus_out)
        self.entries.append(montant_entry)
        self.montant_entry = montant_entry

        self.category_combobox = ttk.Combobox(
            self.entry_frame,
            textvariable=self.category_var,
            values=CATEGORY_VALUES,
            state="readonly",
            style=self.combobox_style,
            exportselection=False,
        )
        self.category_combobox.grid(row=1, column=3, sticky="new", padx=10, pady=(0, 20))
        self.category_combobox.configure(font=self.entry_font)
        self.category_combobox.bind("<<ComboboxSelected>>", self._clear_category_selection)
        self.category_combobox.bind("<FocusOut>", self._clear_category_selection)
        # Ensure dropdown list (popdown) matches theme when opened
        try:
            self._bind_combobox_popdown_theming(self.category_combobox)
        except Exception:
            pass

        # Prélèvement toggle container
        self.prelevement_container = Frame(self.entry_frame, bg=LIGHT_THEME["bg"])
        self.prelevement_container.grid(row=2, column=3, pady=(2, 0))

        # Add "Prélèvement" label
        prelevement_text_label = Label(
            self.prelevement_container,
            text="Prélèvement",
            bg=LIGHT_THEME["bg"],
            fg=LIGHT_THEME["fg"],
            font=self.label_font,
        )
        prelevement_text_label.pack(side="top", pady=(0, 4))
        self.labels.append(prelevement_text_label)

        # Create a sub-container for the toggle elements
        self.prelevement_toggle_container = Frame(self.prelevement_container, bg=LIGHT_THEME["bg"])
        self.prelevement_toggle_container.pack()

        self.prelevement_label = Label(
            self.prelevement_toggle_container,
            text="Non",
            bg=LIGHT_THEME["bg"],
            fg=LIGHT_THEME["fg"],
            font=self.label_font,
        )
        self.prelevement_label.pack(side="left", padx=(0, 6))
        self.prelevement_label.bind("<Button-1>", self._on_prelevement_toggle)
        self.labels.append(self.prelevement_label)

        self.prelevement_canvas = Canvas(
            self.prelevement_toggle_container,
            width=54,
            height=28,
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=LIGHT_THEME["bg"],
        )
        self.prelevement_canvas.pack(side="left", padx=(8, 0))
        self.prelevement_canvas.bind("<Button-1>", self._on_prelevement_toggle)

        self.transaction_container = Frame(self.entry_frame, bg=LIGHT_THEME["bg"])
        self.transaction_container.grid(row=2, column=2, pady=(2, 0))

        self.transaction_label = Label(
            self.transaction_container,
            text="Sortie",
            bg=LIGHT_THEME["bg"],
            fg=LIGHT_THEME["fg"],
            font=self.label_font,
        )
        self.transaction_label.pack(side="left", padx=(0, 6))
        self.transaction_label.bind("<Button-1>", self._on_transaction_toggle)
        self.labels.append(self.transaction_label)

        self.transaction_canvas = Canvas(
            self.transaction_container,
            width=54,
            height=28,
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=LIGHT_THEME["bg"],
        )
        self.transaction_canvas.pack(side="left", padx=(8, 0))
        self.transaction_canvas.bind("<Button-1>", self._on_transaction_toggle)

        # Add the "Ajouter la ligne" button
        self.add_button = Button(
            self.entry_frame,
            text="Ajouter la ligne",
            command=self._add_row,
            relief="flat",
            borderwidth=0,
            padx=20,
            pady=8,
            bg=LIGHT_THEME["toggle_track_active"],
            fg=LIGHT_THEME["toggle_thumb"],
            activebackground=LIGHT_THEME["toggle_track"],
            font=self.label_font,
            cursor="hand2"
        )
        self.add_button.grid(row=3, column=0, columnspan=4, pady=(20, 0))
        # Focus Libellé by default when window opens
        try:
            self.root.after(50, lambda: (self.libelle_entry.focus_set(), self.libelle_entry.icursor('end')))
        except Exception:
            pass
        # --- Path display & change button section (responsive) ---
        self._default_base_parent = Path(__file__).resolve().parent
        self._chosen_parent_dir = self._default_base_parent
        self._compta_folder_name = "dossier_compta"
        # Settings file path & first-run flag
        self._settings_path = Path(__file__).resolve().parent / "settings.json"
        self._first_run = False
        # Accounts (lists)
        self.current_accounts = []  # list[str]
        self.savings_accounts = []  # list[str]
        self.selected_current_account = None  # str | None
        self.selected_savings_account = None  # str | None (for Épargne transfers)
        self._current_account_var = StringVar()
        self._savings_account_var = StringVar()
        # Load persisted settings BEFORE computing base dir / building path UI
        try:
            self._load_settings()
        except Exception:
            self._first_run = True
        # Remove legacy default 'Épargne Principale' if present so savings list starts empty
        if any(acc.strip().lower() == "épargne principale".lower() for acc in self.savings_accounts):
            self.savings_accounts = [acc for acc in self.savings_accounts if acc.strip().lower() != "épargne principale".lower()]
            try:
                self._save_settings()
            except Exception:
                pass
        # Compute effective base directory AFTER potential settings load modifications
        self._compta_base_dir = self._compute_compta_base_dir()
        # Store full path text for dynamic ellipsis rendering
        self._compta_path_full_text = str(self._compta_base_dir)

        self.path_container = Frame(self.entry_frame, bg=LIGHT_THEME["bg"])
        self.path_container.grid(row=4, column=0, columnspan=4, sticky="ew", pady=(14, 0))
        for i in range(3):
            self.path_container.grid_columnconfigure(i, weight=0)
        # Hint label (goes above in narrow mode, inline when wide)
        self.compta_path_hint = Label(
            self.path_container,
            text="Emplacement du dossier compta (Excel):",
            anchor="center",
            bg=LIGHT_THEME["bg"],
            fg=LIGHT_THEME["fg"],
            font=(FONT_FAMILY, max(FONT_MIN_SIZE-1, 9)),
            padx=4,
        )
        self.labels.append(self.compta_path_hint)

        self.compta_path_label = Label(
            self.path_container,
            text=self._compta_path_full_text,
            anchor="center",
            justify="center",
            bg=LIGHT_THEME["bg"],
            fg=LIGHT_THEME["fg"],
            font=self.label_font,
            wraplength=0,
            padx=6,
        )
        self.compta_path_label.bind("<Configure>", lambda e: self._update_path_layout())
        self.labels.append(self.compta_path_label)

        self.change_path_button = Button(
            self.path_container,
            text="Changer...",
            command=self._choose_compta_directory,
            relief="flat",
            borderwidth=0,
            padx=8,
            pady=3,
            bg=LIGHT_THEME["entry_bg"],
            fg=LIGHT_THEME["entry_fg"],
            activebackground=LIGHT_THEME["toggle_track"],
            activeforeground=LIGHT_THEME["toggle_thumb"],
            font=self.small_button_font,
            cursor="hand2"
        )

        # New primary button to open the compta folder in the file explorer
        self.open_path_button = Button(
            self.path_container,
            text="Ouvrir",
            command=self._open_compta_directory,
            relief="flat",
            borderwidth=0,
            padx=8,
            pady=3,
            bg=LIGHT_THEME["entry_bg"],
            fg=LIGHT_THEME["entry_fg"],
            activebackground=LIGHT_THEME["toggle_track"],
            activeforeground=LIGHT_THEME["toggle_thumb"],
            font=self.small_button_font,
            cursor="hand2",
        )
        # Ensure both buttons have exactly the same visual width
        try:
            self._sync_path_buttons_width()
        except Exception:
            pass

        # Initial layout (will be refined in _update_path_layout)
        self.compta_path_hint.grid(row=0, column=0, columnspan=3, pady=(0,4), sticky="n")
        self.compta_path_label.grid(row=1, column=0, columnspan=2, sticky="ew")
        self.open_path_button.grid(row=1, column=2, padx=(12, 4), sticky="e")
        self.change_path_button.grid(row=2, column=2, padx=(12, 4), sticky="e")
        self.path_container.grid_columnconfigure(0, weight=1)
        self.path_container.grid_columnconfigure(1, weight=0)
        self.path_container.grid_columnconfigure(2, weight=0)

        # Final initialization steps
        self.apply_theme()
        self._update_fonts()
        self._clear_category_selection()
        # Draw toggles using the active theme (handles dark on launch)
        self._render_transaction_toggle(self.current_theme)
        self._render_prelevement_toggle(self.current_theme)
        self._initialize_libelle_placeholder()
        self._initialize_montant_placeholder()
        # Inject account selectors after initial theme & fonts
        self._inject_account_selectors()

        self.root.bind("<Configure>", self._on_resize)
        # Bind click events to close calendar when clicking inside the app window
        self.root.bind("<Button-1>", self._maybe_close_calendar, add="+")


    def apply_theme(self):
        theme = DARK_THEME if self.is_dark_mode else LIGHT_THEME
        self.current_theme = theme

        self.root.configure(bg=theme["bg"])
        self.header_frame.configure(bg=theme["bg"])
        self.body.configure(bg=theme["bg"])
        self.entry_frame.configure(bg=theme["bg"])

        # Ensure path section respects theme
        if hasattr(self, "path_container") and self.path_container is not None:
            self.path_container.configure(bg=theme["bg"])
        if hasattr(self, "compta_path_hint") and self.compta_path_hint is not None:
            self.compta_path_hint.configure(bg=theme["bg"], fg=theme["fg"])
        if hasattr(self, "compta_path_label") and self.compta_path_label is not None:
            self.compta_path_label.configure(bg=theme["bg"], fg=theme["fg"])
        if hasattr(self, "change_path_button") and self.change_path_button is not None:
            self.change_path_button.configure(
                bg=theme["entry_bg"],
                fg=theme["entry_fg"],
                activebackground=theme["toggle_track"],
                activeforeground=theme["toggle_thumb"],
            )
        if hasattr(self, "open_path_button") and self.open_path_button is not None:
            self.open_path_button.configure(
                bg=theme["entry_bg"],
                fg=theme["entry_fg"],
                activebackground=theme["toggle_track"],
                activeforeground=theme["toggle_thumb"],
            )

        # Header burger button
        if hasattr(self, "menu_button") and self.menu_button is not None:
            self.menu_button.configure(bg=theme["bg"], fg=theme["fg"], activebackground=theme["entry_bg"], activeforeground=theme["fg"])

        for label in self.labels:
            label.configure(bg=theme["bg"], fg=theme["fg"])

        for entry in self.entries:
            if entry == getattr(self, 'libelle_entry', None):
                entry.configure(bg=theme["entry_bg"], insertbackground=theme["entry_fg"], highlightbackground=theme["entry_border"], highlightcolor=theme["entry_border"])
            elif entry == getattr(self, 'montant_entry', None):
                entry.configure(bg=theme["entry_bg"], insertbackground=theme["entry_fg"], highlightbackground=theme["entry_border"], highlightcolor=theme["entry_border"])
            else:
                entry.configure(bg=theme["entry_bg"], fg=theme["entry_fg"], highlightbackground=theme["entry_border"], highlightcolor=theme["entry_border"])

        # Combobox theming (colors)
        try:
            self.style.configure(
                self.combobox_style,
                fieldbackground=theme["entry_bg"],
                foreground=theme["entry_fg"],
                background=theme["entry_bg"],
                readonlybackground=theme["entry_bg"],
                arrowcolor=theme["fg"],
                bordercolor=theme["entry_border"],
                lightcolor=theme["entry_border"],
                darkcolor=theme["entry_border"],
            )
            # Apply to base style as well (some themes use TCombobox directly for field rendering)
            self.style.configure(
                'TCombobox',
                fieldbackground=theme["entry_bg"],
                foreground=theme["entry_fg"],
                background=theme["entry_bg"],
                readonlybackground=theme["entry_bg"],
                arrowcolor=theme["fg"],
            )
            # Header compact combobox uses same colors but will have tighter padding via _update_fonts
            self.style.configure(
                self.header_combobox_style,
                fieldbackground=theme["entry_bg"],
                foreground=theme["entry_fg"],
                background=theme["entry_bg"],
                readonlybackground=theme["entry_bg"],
                arrowcolor=theme["fg"],
                bordercolor=theme["entry_border"],
                lightcolor=theme["entry_border"],
                darkcolor=theme["entry_border"],
            )
            # Ensure readonly state uses themed colors (Windows/clam)
            self.style.map(
                self.combobox_style,
                fieldbackground=[('readonly', theme["entry_bg"]), ('!readonly', theme["entry_bg"]), ('active', theme["entry_bg"])],
                background=[('readonly', theme["entry_bg"]), ('!readonly', theme["entry_bg"]), ('active', theme["entry_bg"])],
                foreground=[('readonly', theme["entry_fg"]), ('!readonly', theme["entry_fg"]), ('active', theme["entry_fg"])],
            )
            self.style.map(
                'TCombobox',
                fieldbackground=[('readonly', theme["entry_bg"]), ('!readonly', theme["entry_bg"]), ('active', theme["entry_bg"])],
                background=[('readonly', theme["entry_bg"]), ('!readonly', theme["entry_bg"]), ('active', theme["entry_bg"])],
                foreground=[('readonly', theme["entry_fg"]), ('!readonly', theme["entry_fg"]), ('active', theme["entry_fg"])],
            )
            self.style.map(
                self.header_combobox_style,
                fieldbackground=[('readonly', theme["entry_bg"]), ('!readonly', theme["entry_bg"]), ('active', theme["entry_bg"])],
                background=[('readonly', theme["entry_bg"]), ('!readonly', theme["entry_bg"]), ('active', theme["entry_bg"])],
                foreground=[('readonly', theme["entry_fg"]), ('!readonly', theme["entry_fg"]), ('active', theme["entry_fg"])],
            )
        except Exception:
            pass

        # Dropdown list (popdown) colors via option database
        try:
            self.root.option_add('*TCombobox*Listbox.background', theme["entry_bg"])
            self.root.option_add('*TCombobox*Listbox.foreground', theme["entry_fg"])
            self.root.option_add('*TCombobox*Listbox.selectBackground', theme["toggle_track_active"])  # accent
            self.root.option_add('*TCombobox*Listbox.selectForeground', theme["fg"])  # readable on accent
            # Fallback patterns used by some Tk builds
            self.root.option_add('*Combobox*Listbox.background', theme["entry_bg"])
            self.root.option_add('*Combobox*Listbox.foreground', theme["entry_fg"])
            self.root.option_add('*Combobox*Listbox.selectBackground', theme["toggle_track_active"])  # accent
            self.root.option_add('*Combobox*Listbox.selectForeground', theme["fg"])  # readable on accent
            # Generic Listbox defaults as a last resort (affects other listboxes too)
            self.root.option_add('*Listbox.background', theme["entry_bg"])
            self.root.option_add('*Listbox.foreground', theme["entry_fg"])
            self.root.option_add('*Listbox.selectBackground', theme["toggle_track_active"])  # accent
            self.root.option_add('*Listbox.selectForeground', theme["fg"])  # readable on accent
        except Exception:
            pass

        # Re-apply popdown list theming for any open dropdowns
        try:
            if self.category_combobox is not None:
                self._apply_combobox_list_theme(self.category_combobox)
            if hasattr(self, '_current_account_cb') and self._current_account_cb:
                self._apply_combobox_list_theme(self._current_account_cb)
            if hasattr(self, '_savings_account_cb') and self._savings_account_cb:
                self._apply_combobox_list_theme(self._savings_account_cb)
        except Exception:
            pass

        # Update appearance after theme change
        if hasattr(self, 'libelle_entry'):
            self._update_libelle_appearance()
        if hasattr(self, 'montant_entry'):
            self._update_montant_appearance()

        # Apply theme to add button
        def _add_row(self):
            """Handle adding a new row with the current form data"""
            if not EXCEL_AVAILABLE:
                print("Excel functionality not available. Please install openpyxl: pip install openpyxl")
                return
            date_value = self.date_var.get().strip()
            libelle_value = self.libelle_var.get().strip()
            montant_value = self.montant_var.get().strip()
            category_value = self.category_var.get().strip()
            transaction_type = "Entrée" if self.transaction_is_entry else "Sortie"
            prelevement_status = "Oui" if self.is_prelevement else "Non"
            current_account = self.selected_current_account or (self.current_accounts[0] if self.current_accounts else "Compte Principal")
            savings_account = None
            if category_value == "Épargne":
                savings_account = self._savings_account_var.get().strip() or None
                if not savings_account and self.savings_accounts:
                    print("Veuillez sélectionner un compte Épargne.")
                    return
            # Validation
            if not date_value:
                print("Date is required"); return
            if not libelle_value or libelle_value == self.libelle_placeholder:
                print("Libellé is required"); return
            if not montant_value or montant_value == self.montant_placeholder:
                print("Montant is required"); return
            if not category_value:
                print("Catégorie is required"); return
            try:
                parsed_date = datetime.strptime(date_value, "%d-%m-%Y")
                year = parsed_date.year
                month = parsed_date.month
                self._write_to_excel(date_value, libelle_value, montant_value, category_value, transaction_type, prelevement_status, year, month)
                print(f"Successfully added: Date={date_value}, Libellé={libelle_value}, Montant={montant_value}, Catégorie={category_value}, Type={transaction_type}, Prélèvement={prelevement_status}")
                self.libelle_var.set(self.libelle_placeholder); self.libelle_has_placeholder = True; self._update_libelle_appearance()
                self.montant_var.set(self.montant_placeholder); self.montant_has_placeholder = True; self._update_montant_appearance()
                self.category_var.set("")
                try:
                    self.libelle_entry.focus_set(); self.libelle_entry.icursor('end')
                except Exception:
                    pass
                self._clear_category_selection()
                self.transaction_is_entry = False; self.is_prelevement = False; self._savings_account_var.set("")
                self._render_transaction_toggle(self.current_theme); self._render_prelevement_toggle(self.current_theme)
            except ValueError:
                print("Invalid date format. Please use DD-MM-YYYY format.")
            except Exception as e:
                print(f"Error adding row: {e}")


        if self.transaction_container is not None:
            self.transaction_container.configure(bg=theme["bg"])
        if hasattr(self, "transaction_canvas"):
            self.transaction_canvas.configure(bg=theme["bg"])

        if self.prelevement_container is not None:
            self.prelevement_container.configure(bg=theme["bg"])
        if hasattr(self, "prelevement_toggle_container"):
            self.prelevement_toggle_container.configure(bg=theme["bg"])
        if hasattr(self, "prelevement_canvas"):
            self.prelevement_canvas.configure(bg=theme["bg"])

        self._render_transaction_toggle(theme)
        self._render_prelevement_toggle(theme)
        self._refresh_calendar_theme()

        # Theme overlay styling (if created)
        self._apply_menu_theme()

    # ------------------ Overlay Menu ------------------
    def _ensure_menu_overlay(self):
        if self.menu_overlay and self.menu_overlay.winfo_exists():
            return
        theme = self.current_theme
        self.menu_overlay = Frame(self.root, bg=theme["bg"])  # full-size overlay
        # Close button (top-right)
        self.menu_close_button = Button(
            self.menu_overlay,
            text="✕",
            command=self._close_menu,
            relief="flat",
            borderwidth=0,
            bg=theme["bg"],
            fg=theme["fg"],
            activebackground=theme["entry_bg"],
            activeforeground=theme["fg"],
            font=(FONT_FAMILY, 16, "bold"),
            cursor="hand2",
        )
        self.menu_close_button.place(relx=1.0, rely=0.0, x=-12, y=8, anchor="ne")

        # Center content
        self.menu_center_frame = Frame(self.menu_overlay, bg=theme["bg"])
        self.menu_center_frame.place(relx=0.5, rely=0.5, anchor="center")

        self.menu_title_label = Label(
            self.menu_center_frame,
            text="Thème",
            bg=theme["bg"],
            fg=theme["fg"],
            font=(FONT_FAMILY, 18, "bold"),
        )
        self.menu_title_label.pack(pady=(0, 16))

        # Single theme toggle button (sun in dark mode to switch to light, moon in light mode)
        # No background behind the logo: match overlay bg and remove highlights
        self.menu_theme_button = Button(
            self.menu_center_frame,
            image=self._menu_theme_image(),
            relief="flat",
            borderwidth=0,
            highlightthickness=0,
            highlightbackground=theme["bg"],
            bg=theme["bg"],
            activebackground=theme["bg"],
            command=self.toggle_theme,
            cursor="hand2",
        )
        self.menu_theme_button.pack(pady=(0, 4))
        # Directly show accounts editor without intermediate button
        try:
            self._show_accounts_editor()
        except Exception:
            pass

    def _apply_menu_theme(self):
        if not (self.menu_overlay and self.menu_overlay.winfo_exists()):
            return
        theme = self.current_theme
        try:
            self.menu_overlay.configure(bg=theme["bg"])
            self.menu_close_button.configure(bg=theme["bg"], fg=theme["fg"], activebackground=theme["entry_bg"], activeforeground=theme["fg"])
            self.menu_center_frame.configure(bg=theme["bg"])
            self.menu_title_label.configure(bg=theme["bg"], fg=theme["fg"])
            if self.menu_theme_button is not None:
                self.menu_theme_button.configure(
                    bg=theme["bg"],
                    activebackground=theme["bg"],
                    highlightthickness=0,
                    highlightbackground=theme["bg"],
                    image=self._menu_theme_image(),
                )
            # Inline accounts editor theming (recursive)
            if hasattr(self, '_accounts_editor_frame') and self._accounts_editor_frame and self._accounts_editor_frame.winfo_exists():
                try:
                    root_frame = self._accounts_editor_frame
                    root_frame.configure(bg=theme["bg"])
                    stack = [root_frame]
                    while stack:
                        node = stack.pop()
                        try:
                            for child in node.winfo_children():
                                stack.append(child)
                                cls = child.winfo_class()
                                if cls in ('Frame','TFrame'):
                                    try: child.configure(bg=theme["bg"]) ;
                                    except Exception: pass
                                elif cls in ('Label','TLabel'):
                                    try: child.configure(bg=theme["bg"], fg=theme["fg"]) ;
                                    except Exception: pass
                                elif cls in ('Button','TButton'):
                                    try: child.configure(bg=theme["entry_bg"], fg=theme["entry_fg"], activebackground=theme["toggle_track"], activeforeground=theme["toggle_thumb"]) ;
                                    except Exception: pass
                                elif cls == 'Listbox':
                                    try: child.configure(bg=theme["entry_bg"], fg=theme["entry_fg"], selectbackground=theme["toggle_track_active"], selectforeground=theme["fg"]) ;
                                    except Exception: pass
                                elif cls in ('Entry','TEntry'):
                                    try: child.configure(bg=theme["entry_bg"], fg=theme["entry_fg"]) ;
                                    except Exception: pass
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception:
            pass

    def _open_menu(self):
        self._ensure_menu_overlay()
        # Show overlay full screen
        self.menu_overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.menu_overlay.lift()
        # Optional: focus overlay and bind Esc to close
        try:
            self.menu_overlay.focus_set()
            self.menu_overlay.bind("<Escape>", lambda e: self._close_menu())
        except Exception:
            pass
        # Ensure icon is correct for current theme when opening
        self._apply_menu_theme()

    def _close_menu(self):
        if self.menu_overlay and self.menu_overlay.winfo_exists():
            self.menu_overlay.place_forget()

    def _set_theme(self, dark: bool):
        self.is_dark_mode = bool(dark)
        self.apply_theme()
        # Persist preference
        try:
            self._save_settings()
        except Exception:
            pass

    def _menu_theme_image(self):
        """Return sun icon when currently dark (to switch to light), else moon icon."""
        return self.light_mode_icon if self.is_dark_mode else self.dark_mode_icon

    def _refresh_accounts_editor_theme(self):
        """Force re-theming of accounts editor (safe if not present)."""
        try:
            self._apply_menu_theme()
        except Exception:
            pass

    # ---- Inline Accounts Editor (Menu) ----
    def _show_accounts_editor(self):
        """Show side-by-side current & savings accounts editor inside menu overlay."""
        if not (self.menu_overlay and self.menu_overlay.winfo_exists()):
            self._ensure_menu_overlay()
        # Remove previous editor
        try:
            if hasattr(self, '_accounts_editor_frame') and self._accounts_editor_frame and self._accounts_editor_frame.winfo_exists():
                self._accounts_editor_frame.destroy()
        except Exception:
            pass
        theme = self.current_theme
        editor = Frame(self.menu_center_frame, bg=theme['bg'])
        editor.pack(pady=(16,0), fill='both', expand=True)
        self._accounts_editor_frame = editor

        title = Label(editor, text="Gestion des Comptes", font=(FONT_FAMILY, 16, 'bold'), bg=theme['bg'], fg=theme['fg'])
        title.pack(pady=(0,8))

        cols = Frame(editor, bg=theme['bg'])
        cols.pack(fill='both', expand=True)

        # Current accounts column
        cur_col = Frame(cols, bg=theme['bg'])
        cur_col.pack(side='left', fill='both', expand=True, padx=(0,8))
        cur_lbl = Label(cur_col, text="Comptes Courants", font=(FONT_FAMILY, 13, 'bold'), bg=theme['bg'], fg=theme['fg'])
        cur_lbl.pack(anchor='w')
        self._menu_cur_listbox = Listbox(cur_col, height=6, activestyle='none')
        self._menu_cur_listbox.pack(fill='x', pady=4)
        for acc in self.current_accounts:
            try: self._menu_cur_listbox.insert('end', acc)
            except Exception: pass
        cur_add_row = Frame(cur_col, bg=theme['bg'])
        cur_add_row.pack(fill='x', pady=(2,4))
        self._menu_cur_new_var = StringVar()
        cur_entry = Entry(cur_add_row, textvariable=self._menu_cur_new_var, width=18)
        cur_entry.pack(side='left', padx=(0,6))
        cur_add_btn = Button(cur_add_row, text="Ajouter", command=self._menu_add_current_account, relief='flat', bg=theme['toggle_track_active'], fg=theme['toggle_thumb'])
        cur_add_btn.pack(side='left')
        cur_remove_btn = Button(cur_col, text="Supprimer sélection", command=self._menu_remove_current_account, relief='flat', bg=theme['entry_bg'], fg=theme['entry_fg'])
        cur_remove_btn.pack(anchor='e', pady=(0,4))

        # Savings accounts column
        sav_col = Frame(cols, bg=theme['bg'])
        sav_col.pack(side='left', fill='both', expand=True, padx=(8,0))
        sav_lbl = Label(sav_col, text="Comptes Épargne", font=(FONT_FAMILY, 13, 'bold'), bg=theme['bg'], fg=theme['fg'])
        sav_lbl.pack(anchor='w')
        self._menu_sav_listbox = Listbox(sav_col, height=6, activestyle='none')
        self._menu_sav_listbox.pack(fill='x', pady=4)
        for acc in self.savings_accounts:
            try: self._menu_sav_listbox.insert('end', acc)
            except Exception: pass
        sav_add_row = Frame(sav_col, bg=theme['bg'])
        sav_add_row.pack(fill='x', pady=(2,4))
        self._menu_sav_new_var = StringVar()
        sav_entry = Entry(sav_add_row, textvariable=self._menu_sav_new_var, width=18)
        sav_entry.pack(side='left', padx=(0,6))
        sav_add_btn = Button(sav_add_row, text="Ajouter", command=self._menu_add_savings_account, relief='flat', bg=theme['toggle_track_active'], fg=theme['toggle_thumb'])
        sav_add_btn.pack(side='left')
        sav_remove_btn = Button(sav_col, text="Supprimer sélection", command=self._menu_remove_savings_account, relief='flat', bg=theme['entry_bg'], fg=theme['entry_fg'])
        sav_remove_btn.pack(anchor='e', pady=(0,4))

        # Save button & feedback
        self._menu_feedback_var = StringVar()
        save_btn = Button(editor, text="Enregistrer", command=self._menu_save_accounts, relief='flat', bg=theme['toggle_track_active'], fg=theme['toggle_thumb'], padx=16, pady=6, font=(FONT_FAMILY, 13, 'bold'))
        save_btn.pack(pady=(12,4))
        feedback = Label(editor, textvariable=self._menu_feedback_var, font=(FONT_FAMILY, 11), bg=theme['bg'], fg=theme['fg'])
        feedback.pack()

        try: cur_entry.focus_set()
        except Exception: pass
        self._apply_menu_theme()

    def _menu_add_current_account(self):
        name = (getattr(self, '_menu_cur_new_var', StringVar()).get() or '').strip()
        if not name or name in self.current_accounts:
            return
        # Clear previous save feedback on modification
        try:
            if hasattr(self, '_menu_feedback_var'):
                self._menu_feedback_var.set('')
        except Exception:
            pass
        self.current_accounts.append(name)
        try: self._menu_cur_listbox.insert('end', name)
        except Exception: pass
        try: self._menu_cur_new_var.set('')
        except Exception: pass

    def _menu_remove_current_account(self):
        try:
            sel = self._menu_cur_listbox.curselection()
            if not sel:
                return
            idx = sel[0]
            acc = self._menu_cur_listbox.get(idx)
            if acc in self.current_accounts:
                self.current_accounts = [a for a in self.current_accounts if a != acc]
            self._menu_cur_listbox.delete(idx)
            try:
                if hasattr(self, '_menu_feedback_var'):
                    self._menu_feedback_var.set('')
            except Exception:
                pass
        except Exception:
            pass

    def _menu_add_savings_account(self):
        name = (getattr(self, '_menu_sav_new_var', StringVar()).get() or '').strip()
        if not name or name in self.savings_accounts:
            return
        try:
            if hasattr(self, '_menu_feedback_var'):
                self._menu_feedback_var.set('')
        except Exception:
            pass
        self.savings_accounts.append(name)
        try: self._menu_sav_listbox.insert('end', name)
        except Exception: pass
        try: self._menu_sav_new_var.set('')
        except Exception: pass

    def _menu_remove_savings_account(self):
        try:
            sel = self._menu_sav_listbox.curselection()
            if not sel:
                return
            idx = sel[0]
            acc = self._menu_sav_listbox.get(idx)
            if acc in self.savings_accounts:
                self.savings_accounts = [a for a in self.savings_accounts if a != acc]
            self._menu_sav_listbox.delete(idx)
            try:
                if hasattr(self, '_menu_feedback_var'):
                    self._menu_feedback_var.set('')
            except Exception:
                pass
        except Exception:
            pass

    def _menu_save_accounts(self):
        if not self.current_accounts:
            self.current_accounts = ['Compte Courant']
        try:
            self._save_settings()
            if hasattr(self, '_menu_feedback_var'):
                self._menu_feedback_var.set('Enregistré ✔')
            try: self._inject_account_selectors()
            except Exception: pass
        except Exception:
            if hasattr(self, '_menu_feedback_var'):
                self._menu_feedback_var.set('Erreur sauvegarde')

    # ----- Combobox popdown theming helpers -----
    def _apply_combobox_list_theme(self, cb):
        """Apply current theme to the dropdown list (popdown) of a ttk.Combobox."""
        theme = self.current_theme
        try:
            # Flush pending geometry so popdown widgets exist
            try:
                self.root.update_idletasks()
            except Exception:
                pass
            popdown = cb.tk.call('ttk::combobox::PopdownWindow', str(cb))
            pop = None
            try:
                pop = cb.nametowidget(popdown)
                pop.configure(bg=theme["entry_bg"])
            except Exception:
                pass

            # Try the standard path first
            def _style_listbox(lb):
                try:
                    lb.configure(
                        background=theme["entry_bg"],
                        bg=theme["entry_bg"],
                        foreground=theme["entry_fg"],
                        fg=theme["entry_fg"],
                        selectbackground=theme["toggle_track_active"],
                        selectforeground=theme["fg"],
                        highlightbackground=theme["entry_bg"],
                        highlightcolor=theme["entry_bg"],
                        highlightthickness=0,
                        relief='flat',
                    )
                    try:
                        # Apply larger dropdown font for readability
                        lb.configure(font=self.dropdown_font)
                    except Exception:
                        pass
                    try:
                        lb.update_idletasks()
                    except Exception:
                        pass
                except Exception:
                    pass

            # Standard child path
            try:
                frame = cb.nametowidget(popdown + '.f')
                try:
                    frame.configure(bg=theme["entry_bg"])
                except Exception:
                    pass
                try:
                    lb = cb.nametowidget(popdown + '.f.l')
                    _style_listbox(lb)
                except Exception:
                    pass
            except Exception:
                pass

            # Fallback: recursively search for any Listbox under the popdown
            try:
                root_w = pop if pop is not None else cb.nametowidget(popdown)
                stack = [root_w]
                while stack:
                    w = stack.pop()
                    try:
                        wclass = w.winfo_class()
                    except Exception:
                        wclass = ''
                    # Theme container backgrounds
                    if wclass in ('TFrame', 'Frame', 'Toplevel'):
                        try:
                            w.configure(bg=theme["entry_bg"])
                        except Exception:
                            pass
                    if wclass == 'Listbox':
                        _style_listbox(w)
                    # Recurse
                    try:
                        stack.extend(w.winfo_children())
                    except Exception:
                        pass
            except Exception:
                # If the listbox isn't ready yet, try again shortly
                try:
                    self.root.after(20, lambda: self._apply_combobox_list_theme(cb))
                except Exception:
                    pass
        except Exception:
            pass

    def _bind_combobox_popdown_theming(self, cb):
        """Bind events so when the combobox opens, its dropdown list gets themed."""
        def _apply_later(_=None):
            # Try a few times after open to handle platform timing
            self._schedule_popdown_theme(cb, attempts=4, delay=15)
        try:
            cb.bind('<Button-1>', _apply_later, add='+')
            cb.bind('<Alt-Down>', _apply_later, add='+')
            cb.bind('<F4>', _apply_later, add='+')
        except Exception:
            pass

    def _schedule_popdown_theme(self, cb, attempts=10, delay=20):
        """Schedule several attempts to apply popdown theming to handle creation timing."""
        try:
            for i in range(max(1, int(attempts))):
                self.root.after(delay * (i + 1), lambda c=cb: self._apply_combobox_list_theme(c))
        except Exception:
            pass

    def _clear_category_selection(self, *_):
        if self.category_combobox is not None:
            self.category_combobox.selection_clear()
            self.category_combobox.icursor("end")

    def _validate_amount(self, proposed: str) -> bool:
        # Allow empty string and placeholder value
        if proposed == "" or proposed == self.montant_placeholder:
            return True
        allowed_chars = set("0123456789.,")
        if any(char not in allowed_chars for char in proposed):
            return False
        if proposed.count(".") + proposed.count(",") > 1:
            return False
        return True

    def _on_montant_focus_in(self, event):
        """Handle focus in event for montant entry - remove placeholder"""
        if self.montant_has_placeholder:
            self.montant_var.set("")
            self.montant_has_placeholder = False
            self._update_montant_appearance()

    def _on_montant_focus_out(self, event):
        """Handle focus out event for montant entry - add placeholder if empty"""
        if not self.montant_var.get().strip():
            self.montant_var.set(self.montant_placeholder)
            self.montant_has_placeholder = True
            self._update_montant_appearance()

    def _update_montant_appearance(self):
        """Update the appearance of montant entry based on placeholder state"""
        if hasattr(self, 'montant_entry') and self.montant_entry:
            theme = self.current_theme
            if self.montant_has_placeholder:
                # Placeholder style - lighter text
                self.montant_entry.configure(
                    fg=theme["entry_border"]  # Use border color for placeholder (lighter)
                )
            else:
                # Normal style
                self.montant_entry.configure(
                    fg=theme["entry_fg"]
                )

    def _initialize_montant_placeholder(self):
        """Initialize the placeholder in the montant entry"""
        if not self.montant_var.get().strip():
            self.montant_var.set(self.montant_placeholder)
            self.montant_has_placeholder = True
            self._update_montant_appearance()

    def _on_libelle_focus_in(self, event):
        """Handle focus in event for libellé entry - remove placeholder"""
        if self.libelle_has_placeholder:
            self.libelle_var.set("")
            self.libelle_has_placeholder = False
            self._update_libelle_appearance()

    def _on_libelle_focus_out(self, event):
        """Handle focus out event for libellé entry - add placeholder if empty"""
        if not self.libelle_var.get().strip():
            self.libelle_var.set(self.libelle_placeholder)
            self.libelle_has_placeholder = True
            self._update_libelle_appearance()

    def _update_libelle_appearance(self):
        """Update the appearance of libellé entry based on placeholder state"""
        if hasattr(self, 'libelle_entry') and self.libelle_entry:
            theme = self.current_theme
            if self.libelle_has_placeholder:
                # Placeholder style - lighter text
                self.libelle_entry.configure(
                    fg=theme["entry_border"]  # Use border color for placeholder (lighter)
                )
            else:
                # Normal style
                self.libelle_entry.configure(
                    fg=theme["entry_fg"]
                )

    def _initialize_libelle_placeholder(self):
        """Initialize the placeholder in the libellé entry"""
        if not self.libelle_var.get().strip():
            self.libelle_var.set(self.libelle_placeholder)
            self.libelle_has_placeholder = True
            self._update_libelle_appearance()

    def _render_transaction_toggle(self, theme):
        if not hasattr(self, "transaction_canvas"):
            return
        self.transaction_canvas.delete("all")
        self.transaction_canvas.configure(bg=theme["bg"])
        track_color = theme["toggle_track_active"] if self.transaction_is_entry else theme["toggle_track"]
        thumb_color = theme["toggle_thumb"]

        width = int(self.transaction_canvas.cget("width"))
        height = int(self.transaction_canvas.cget("height"))
        radius = height // 2

        self.transaction_canvas.create_oval(0, 0, height, height, fill=track_color, outline="")
        self.transaction_canvas.create_oval(width - height, 0, width, height, fill=track_color, outline="")
        self.transaction_canvas.create_rectangle(radius, 0, width - radius, height, fill=track_color, outline="")

        thumb_offset = width - height + 2 if self.transaction_is_entry else 2
        self.transaction_canvas.create_oval(
            thumb_offset,
            2,
            thumb_offset + height - 4,
            height - 2,
            fill=thumb_color,
            outline="",
        )

        self.transaction_label.configure(text="Entrée" if self.transaction_is_entry else "Sortie")

    def _on_transaction_toggle(self, *_):
        self.transaction_is_entry = not self.transaction_is_entry
        self._render_transaction_toggle(self.current_theme)

    def _render_prelevement_toggle(self, theme):
        if not hasattr(self, "prelevement_canvas"):
            return
        self.prelevement_canvas.delete("all")
        self.prelevement_canvas.configure(bg=theme["bg"])
        track_color = theme["toggle_track_active"] if self.is_prelevement else theme["toggle_track"]
        thumb_color = theme["toggle_thumb"]

        width = int(self.prelevement_canvas.cget("width"))
        height = int(self.prelevement_canvas.cget("height"))
        radius = height // 2

        self.prelevement_canvas.create_oval(0, 0, height, height, fill=track_color, outline="")
        self.prelevement_canvas.create_oval(width - height, 0, width, height, fill=track_color, outline="")
        self.prelevement_canvas.create_rectangle(radius, 0, width - radius, height, fill=track_color, outline="")

        thumb_offset = width - height + 2 if self.is_prelevement else 2
        self.prelevement_canvas.create_oval(
            thumb_offset,
            2,
            thumb_offset + height - 4,
            height - 2,
            fill=thumb_color,
            outline="",
        )

        self.prelevement_label.configure(text="Oui" if self.is_prelevement else "Non")

    def _on_prelevement_toggle(self, *_):
        self.is_prelevement = not self.is_prelevement
        self._render_prelevement_toggle(self.current_theme)

    def _on_resize(self, event):
        if event.widget is self.root:
            self._update_fonts()
            # Update responsive path layout after fonts/size change
            try:
                self._update_path_layout()
            except Exception:
                pass

    def _update_fonts(self):
        width = max(self.root.winfo_width(), self.root.winfo_reqwidth())
        height = max(self.root.winfo_height(), self.root.winfo_reqheight())
        reference = min(width, height)
        new_size = max(FONT_MIN_SIZE, min(FONT_MAX_SIZE, reference // FONT_SCALE_DIVISOR))

        if self.entry_font.cget("size") == new_size:
            return

        self.entry_font.configure(size=new_size)
        self.label_font.configure(size=new_size)
        # Header combobox font now matches entry font for uniform appearance
        header_size = new_size
        self.header_cb_font.configure(size=header_size)

        for label in self.labels:
            label.configure(font=self.label_font)
        for entry in self.entries:
            entry.configure(font=self.entry_font)

        # Update add button font
        if hasattr(self, 'add_button'):
            self.add_button.configure(font=self.label_font)
        if hasattr(self, 'change_path_button'):
            try:
                self.small_button_font.configure(size=max(7, new_size - 4))
            except Exception:
                pass
            self.change_path_button.configure(font=self.small_button_font)
        if hasattr(self, 'open_path_button'):
            self.open_path_button.configure(font=self.small_button_font)
        # Keep both path buttons with identical width when fonts change
        try:
            self._sync_path_buttons_width()
        except Exception:
            pass

        if self.category_combobox is not None:
            self.category_combobox.configure(font=self.entry_font)
            # Style: scale combobox font, padding, and arrow size to match overall UI
            try:
                # Font
                self.style.configure(self.combobox_style, font=self.entry_font)
                # Slightly smaller padding/arrow for a leaner look
                linespace = max(1, int(self.entry_font.metrics('linespace')))
                pad_y = max(2, int(linespace * 0.20))
                arrowsize = max(10, int(linespace * 0.80))
                self.style.configure(self.combobox_style, padding=(6, pad_y), arrowsize=arrowsize)
                # Even tighter padding/arrow for header compact style
                # Remove special compact header style; use unified style for all dropdowns
                self.style.configure(self.header_combobox_style, font=self.header_cb_font, padding=(6, pad_y), arrowsize=arrowsize)
            except Exception:
                pass
            # Ensure other comboboxes follow the same font (account/savings selectors)
            try:
                if hasattr(self, '_current_account_cb') and self._current_account_cb:
                    # Use unified header font (same as entry) and standard vertical padding
                    self._current_account_cb.configure(font=self.header_cb_font)
                    try:
                        self._current_account_cb.pack_configure(pady=pad_y)
                    except Exception:
                        pass
                if hasattr(self, '_savings_account_cb') and self._savings_account_cb:
                    self._savings_account_cb.configure(font=self.entry_font)
            except Exception:
                pass
            # Configure dropdown list (popdown) font via option database
            try:
                # Increase dropdown list item font size for better readability
                try:
                    # Keep dropdown list a bit smaller than the combobox field text
                    dd_size = max(8, int(self.entry_font.cget('size')) - 4)
                    if self.dropdown_font.cget('size') != dd_size:
                        self.dropdown_font.configure(size=dd_size)
                except Exception:
                    pass
                self.root.option_add('*TCombobox*Listbox.font', self.dropdown_font)
            except Exception:
                pass
            self._clear_category_selection()
        # Dynamic toggle sizing based on computed font size / window
        toggle_height = int(new_size * 1.7)
        toggle_height = max(TOGGLE_MIN_HEIGHT, min(TOGGLE_MAX_HEIGHT, toggle_height))
        toggle_width = int(toggle_height * TOGGLE_ASPECT_RATIO)

        resized = False
        if hasattr(self, 'transaction_canvas') and self.transaction_canvas:
            cur_w = int(self.transaction_canvas.cget('width'))
            cur_h = int(self.transaction_canvas.cget('height'))
            if cur_w != toggle_width or cur_h != toggle_height:
                self.transaction_canvas.configure(width=toggle_width, height=toggle_height)
                resized = True
        if hasattr(self, 'prelevement_canvas') and self.prelevement_canvas:
            cur_w = int(self.prelevement_canvas.cget('width'))
            cur_h = int(self.prelevement_canvas.cget('height'))
            if cur_w != toggle_width or cur_h != toggle_height:
                self.prelevement_canvas.configure(width=toggle_width, height=toggle_height)
                resized = True

        # Re-render toggles (always if resized; cheap operation)
        if resized:
            self._render_transaction_toggle(self.current_theme)
            self._render_prelevement_toggle(self.current_theme)
        else:
            self._render_transaction_toggle(self.current_theme)
            self._render_prelevement_toggle(self.current_theme)
        if self.calendar_win is not None:
            self._build_calendar_body()
        # Also adjust path layout once sizes settle
        try:
            self._update_path_layout()
        except Exception:
            pass

    def _update_path_layout(self):
        """Responsive layout for path section: stacked (hint above) when narrow, inline when wide."""
        if not hasattr(self, 'path_container'):
            return
        cont = self.path_container
        width = cont.winfo_width()
        if width <= 1:  # Not yet rendered; schedule later
            self.root.after(50, self._update_path_layout)
            return
        # Clear current grid placements
        try:
            self.compta_path_hint.grid_forget()
            self.compta_path_label.grid_forget()
            if hasattr(self, 'open_path_button'):
                self.open_path_button.grid_forget()
            if hasattr(self, 'change_path_button'):
                self.change_path_button.grid_forget()
        except Exception:
            pass
        threshold = 900
        # Wide mode: all in one line
        if width >= threshold:
            for i in range(3):
                cont.grid_columnconfigure(i, weight=0)
            cont.grid_columnconfigure(1, weight=1)
            self.compta_path_hint.configure(anchor='center')
            # Standard hint padding; overall vertical spacing handled on container
            self.compta_path_hint.grid(row=0, column=0, padx=(4,8), pady=(0,0), sticky='w')
            # Prevent wrapping; we'll show ellipsis when too narrow
            self.compta_path_label.configure(wraplength=0, anchor='center', justify='center')
            self.compta_path_label.grid(row=0, column=1, padx=(0,12), sticky='ew')
            # Place Ouvrir as primary on the right, with Changer below it
            if hasattr(self, 'open_path_button'):
                self.open_path_button.grid(row=0, column=2, padx=(0,6), sticky='e')
            if hasattr(self, 'change_path_button'):
                self.change_path_button.grid(row=1, column=2, padx=(0,6), pady=(4,0), sticky='e')
            # Apply large spacing above the entire path container once (distance from add button)
            try:
                self.path_container.grid_configure(pady=(100,0))
            except Exception:
                pass
        else:
            # Stacked mode
            cont.grid_columnconfigure(0, weight=1)
            cont.grid_columnconfigure(1, weight=0)
            cont.grid_columnconfigure(2, weight=0)
            self.compta_path_hint.configure(anchor='center')
            self.compta_path_hint.grid(row=0, column=0, columnspan=3, pady=(0,6))
            # Prevent wrapping; we'll show ellipsis when too narrow
            self.compta_path_label.configure(wraplength=0, anchor='center', justify='center')
            self.compta_path_label.grid(row=1, column=0, columnspan=2, sticky='ew')
            # Ouvrir on the right of the label row, Changer below it
            if hasattr(self, 'open_path_button'):
                self.open_path_button.grid(row=1, column=2, padx=(12,4), sticky='e')
            if hasattr(self, 'change_path_button'):
                self.change_path_button.grid(row=2, column=2, padx=(12,4), pady=(4,0), sticky='e')
            # Restore default smaller container spacing in stacked layout
            try:
                self.path_container.grid_configure(pady=(14,0))
            except Exception:
                pass
        # Force geometry update for smoother transitions
        try:
            cont.update_idletasks()
        except Exception:
            pass
        # After layout is updated, ensure the path text is elided to fit
        try:
            self._refresh_compta_path_elision()
        except Exception:
            pass

    def _elide_text_to_width(self, text: str, max_width_px: int) -> str:
        """Return text truncated with '...' so that it fits within max_width_px using label font.

        Ellipsis is added at the end (right truncation). If even '...' doesn't fit, returns '...'.
        """
        try:
            fnt = self.label_font
            if fnt.measure(text) <= max_width_px:
                return text
            ell = '...'
            if fnt.measure(ell) > max_width_px:
                return '...'
            lo, hi = 0, len(text)
            # Binary search for max prefix length that fits with ellipsis
            while lo < hi:
                mid = (lo + hi) // 2
                candidate = text[:mid] + ell
                if fnt.measure(candidate) <= max_width_px:
                    lo = mid + 1
                else:
                    hi = mid
            cut = max(0, hi - 1)
            return (text[:cut] + ell) if cut > 0 else '...'
        except Exception:
            # Fallback to simple char-based truncation
            if len(text) <= 70:
                return text
            return text[:67] + '...'

    def _refresh_compta_path_elision(self):
        """Update the displayed path with an ellipsis based on the label's current width."""
        if not hasattr(self, 'compta_path_label'):
            return
        full_text = getattr(self, '_compta_path_full_text', str(self._compta_base_dir))
        lbl = self.compta_path_label
        try:
            lbl.update_idletasks()
        except Exception:
            pass
        width = lbl.winfo_width()
        if width <= 1:
            # Try again shortly when geometry is known
            self.root.after(50, self._refresh_compta_path_elision)
            return
        # Account for internal horizontal padding a bit
        pad = 8
        max_px = max(0, width - pad)
        new_text = self._elide_text_to_width(full_text, max_px)
        # Avoid unnecessary updates
        if lbl.cget('text') != new_text:
            lbl.configure(text=new_text)

    def _sync_path_buttons_width(self):
        """Make 'Ouvrir' and 'Changer...' buttons have identical visual widths.

        Width is set in "character units" based on the current label font so it stays consistent
        as the UI scales. This should be called after creating the buttons and whenever fonts change.
        """
        try:
            if not hasattr(self, 'open_path_button') or not hasattr(self, 'change_path_button'):
                return
            if self.open_path_button is None or self.change_path_button is None:
                return
            fnt = self.label_font
            # Measure text widths
            text_open = str(self.open_path_button.cget('text') or '')
            text_change = str(self.change_path_button.cget('text') or '')
            max_text_px = max(fnt.measure(text_open), fnt.measure(text_change))
            # Account for horizontal padding (in pixels) set on the buttons
            def _padx(btn):
                try:
                    return int(str(btn.cget('padx')))
                except Exception:
                    return 0
            pad = max(_padx(self.open_path_button), _padx(self.change_path_button))
            total_px = max_text_px + 2 * pad
            # Convert to character units using width of '0' in this font
            char_px = max(1, fnt.measure('0'))
            width_chars = max(1, (total_px + char_px - 1) // char_px)
            self.open_path_button.configure(width=width_chars)
            self.change_path_button.configure(width=width_chars)
        except Exception:
            pass

    def _add_row(self):
        """Handle adding a new row with the current form data"""
        if not EXCEL_AVAILABLE:
            print("Excel functionality not available. Please install openpyxl: pip install openpyxl")
            return

        # Get the form data
        date_value = self.date_var.get().strip()
        libelle_value = self.libelle_var.get().strip()
        montant_value = self.montant_var.get().strip()
        category_value = self.category_var.get().strip()
        transaction_type = "Entrée" if self.transaction_is_entry else "Sortie"
        prelevement_status = "Oui" if self.is_prelevement else "Non"

        # Basic validation
        if not date_value:
            print("Date is required")
            return
        if not libelle_value or libelle_value == self.libelle_placeholder:
            print("Libellé is required")
            return
        if not montant_value or montant_value == self.montant_placeholder:
            print("Montant is required")
            return
        if not category_value:
            print("Catégorie is required")
            return

        try:
            # Parse the date to get year and month
            parsed_date = datetime.strptime(date_value, "%d-%m-%Y")
            year = parsed_date.year
            month = parsed_date.month

            # Write to Excel
            self._write_to_excel(date_value, libelle_value, montant_value, category_value, transaction_type, prelevement_status, year, month)


            print(f"Successfully added: Date={date_value}, Libellé={libelle_value}, Montant={montant_value}, Catégorie={category_value}, Type={transaction_type}, Prélèvement={prelevement_status}")

            # Clear the form after adding
            self.libelle_var.set(self.libelle_placeholder)
            self.libelle_has_placeholder = True
            self._update_libelle_appearance()
            self.montant_var.set(self.montant_placeholder)
            self.montant_has_placeholder = True
            self._update_montant_appearance()
            self.category_var.set("")
            # Keep the date as it might be reused
            # Return focus to libellé for rapid consecutive entry
            try:
                self.libelle_entry.focus_set()
                self.libelle_entry.icursor('end')
            except Exception:
                pass

            # Clear category selection
            self._clear_category_selection()

            # Reset toggles to default after successful add
            self.transaction_is_entry = False
            self.is_prelevement = False
            self._render_transaction_toggle(self.current_theme)
            self._render_prelevement_toggle(self.current_theme)

        except ValueError:
            print("Invalid date format. Please use DD-MM-YYYY format.")
        except Exception as e:
            print(f"Error adding row: {e}")

    def _write_to_excel(self, date_value, libelle_value, montant_value, category_value, transaction_type, prelevement_status, year, month):
        """Write data to Excel file with the specified structure"""
        if not EXCEL_AVAILABLE:
            return
        # Lazy ensure current base directory exists (user-configurable) ONLY when writing
        base_dir = self._compta_base_dir
        if not base_dir.exists():
            try:
                base_dir.mkdir(parents=True, exist_ok=True)
                print(f"Création du dossier compta: {base_dir}")
                # Refresh label to remove the 'sera créé' hint if present
                if hasattr(self, 'compta_path_label'):
                    self._compta_path_full_text = str(base_dir)
                    self.compta_path_label.configure(text=self._compta_path_full_text)
                    try:
                        self._refresh_compta_path_elision()
                    except Exception:
                        pass
            except Exception as e:
                print(f"Warning: could not create directory {base_dir}: {e}")
        # Create file path inside the subdirectory
        filename = f"Compta_{year}.xlsx"
        file_path = base_dir / filename

        # Sheet name format: MM_YYYY
        sheet_name = f"{month:02d}_{year}"

        # Handle file locks with a simple guided retry flow
        while True:
            try:
                # Try to load existing workbook
                if file_path.exists():
                    wb = load_workbook(file_path)
                else:
                    wb = Workbook()
                    # Remove default sheet
                    if "Sheet" in wb.sheetnames:
                        wb.remove(wb["Sheet"])

                # Get or create the sheet for this month
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
                    # Add headers including Transaction column G
                    headers = ["Date", "Libellé", "Montant", "Catégorie", "Type", "Prélèvement", "Transaction"]
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)

                # Find the next empty row considering transaction data columns (A-G)
                next_row = self._find_next_transaction_row(ws)

                # Add the data
                # Column A: store as real date (not text)
                date_cell = ws.cell(row=next_row, column=1)
                try:
                    parsed_dt = datetime.strptime(date_value, "%d-%m-%Y").date()
                    date_cell.value = parsed_dt
                    # Use localized short date style (day/month/year)
                    date_cell.number_format = "dd/mm/yyyy"
                except Exception:
                    # Fallback to original text if parsing fails
                    date_cell.value = date_value
                ws.cell(row=next_row, column=2, value=libelle_value)
                montant_cell = ws.cell(row=next_row, column=3)
                try:
                    montant_numeric = float(str(montant_value).replace(",", "."))
                    montant_cell.value = montant_numeric
                    montant_cell.number_format = "#,##0.00 [$€-fr-FR]"
                except Exception:
                    montant_cell.value = montant_value
                ws.cell(row=next_row, column=4, value=category_value)
                ws.cell(row=next_row, column=5, value=transaction_type)
                ws.cell(row=next_row, column=6, value=prelevement_status)
                # Column G Transaction: directional string from -> to
                try:
                    current_account = self.selected_current_account or (self.current_accounts[0] if self.current_accounts else "Compte Courant")
                except Exception:
                    current_account = "Compte Courant"
                transaction_str = current_account
                if category_value == "Épargne":
                    savings_account = getattr(self, '_savings_account_var', StringVar()).get().strip()
                    if savings_account:
                        if transaction_type == "Entrée":
                            # Withdrawal from savings into current (savings loses, current gains)
                            transaction_str = f"{savings_account} -> {current_account}"
                        else:
                            # Deposit from current into savings (current loses, savings gains)
                            transaction_str = f"{current_account} -> {savings_account}"
                ws.cell(row=next_row, column=7, value=transaction_str)

                # Apply color formatting with priority: Prélèvement > Transaction Type
                if prelevement_status == "Oui":
                    # Yellow background for prélèvement (light yellow with dark orange text)
                    fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    font_color = Font(color="F57F17")
                elif transaction_type == "Entrée":
                    # Green background for income (light green with dark green text)
                    fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    font_color = Font(color="2E7D32")
                else:  # "Sortie"
                    # Red background for outcome (light red with dark red text)
                    fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                    font_color = Font(color="C62828")

                # Apply formatting to all cells in the row (columns 1-7 now)
                for col in range(1, 8):  # Columns 1-7
                    cell = ws.cell(row=next_row, column=col)
                    cell.fill = fill
                    cell.font = font_color

                # Normalize montant column formats
                try:
                    for r in range(2, ws.max_row + 1):
                        c_amt = ws.cell(row=r, column=3)
                        v_amt = c_amt.value
                        if isinstance(v_amt, str):
                            try:
                                c_amt.value = float(v_amt.replace(",", "."))
                            except Exception:
                                pass
                        if isinstance(c_amt.value, (int, float)):
                            c_amt.number_format = "#,##0.00 [$€-fr-FR]"
                except Exception:
                    pass

                # Try to save the workbook
                # Update statistics for this sheet before saving
                try:
                    self._update_sheet_statistics(ws)
                except Exception as e:
                    print(f"Warning: could not update statistics: {e}")
                # Normalize existing date column values to real dates
                try:
                    self._normalize_date_column(ws)
                except Exception as e:
                    print(f"Warning: could not normalize dates: {e}")
                # Apply fixed column widths (A-F = 20, H-I = 30) before font sizing
                try:
                    self._apply_fixed_widths(ws)
                except Exception as e:
                    print(f"Warning: could not apply fixed widths: {e}")
                # Apply font sizing (headers vs data) before saving
                try:
                    self._apply_sheet_fonts(ws)
                except Exception as e:
                    print(f"Warning: could not apply font sizing: {e}")
                wb.save(file_path)
                try:
                    rel_path = file_path.relative_to(Path(__file__).resolve().parent)
                except Exception:
                    rel_path = file_path
                print(f"Data saved to {rel_path}, sheet {sheet_name}")
                break

            except PermissionError:
                # Excel likely has the file open. Ask user to close and retry.
                retry = messagebox.askretrycancel(
                    title="Excel file is open",
                    message=(
                        f"Le fichier {filename} est ouvert dans Excel et ne peut pas être modifié.\n\n"
                        "Fermez le fichier dans Excel puis cliquez sur Réessayer.\n"
                        "Cliquez sur Annuler pour abandonner l'enregistrement."
                    ),
                )
                if not retry:
                    print("Save cancelled by user due to open Excel file.")
                    return
            except Exception as e:
                print(f"Error writing to Excel: {e}")
                raise

    # ------------------ Compta Directory Handling ------------------
    def _compute_compta_base_dir(self):
        """Return the effective dossier_compta directory under the chosen parent.

        If the chosen parent already ends with the target folder name, use it directly.
        """
        try:
            if self._chosen_parent_dir.name == self._compta_folder_name:
                return self._chosen_parent_dir
            return self._chosen_parent_dir / self._compta_folder_name
        except Exception:
            return Path(__file__).resolve().parent / self._compta_folder_name

    def _format_compta_path_for_label(self, path: Path) -> str:
        """Format path for label display.

        Simple right truncation with '...' suffix when exceeding max length.
        Keeps the beginning (most informative root) and trims the tail.
        """
        s = str(path)
        # Allow external tuning if desired later
        max_len = 70
        if len(s) <= max_len:
            return s
        if max_len <= 3:
            return s[:max_len]
        return s[: max_len - 3] + '...'

    def _choose_compta_directory(self):
        """Open a dialog to let user pick the parent directory for the compta folder."""
        try:
            initial = str(self._chosen_parent_dir)
        except Exception:
            initial = str(Path.home())
        try:
            selected = filedialog.askdirectory(title="Choisir le dossier parent pour 'dossier_compta'", initialdir=initial)
        except Exception as e:
            print(f"Dialog error: {e}")
            return
        if not selected:
            return
        try:
            new_parent = Path(selected)
            # Do NOT create the compta folder or even the chosen parent (unless parent truly missing) yet.
            # Folder creation is deferred until a first line is added (lazy creation).
            if not new_parent.exists():
                # We still need the parent directory itself to exist to avoid user confusion, create parent only.
                try:
                    new_parent.mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    print(f"Impossible de créer le dossier parent sélectionné: {e}")
                    return
            self._chosen_parent_dir = new_parent
            self._compta_base_dir = self._compute_compta_base_dir()
            # Update label to reflect future path (not yet created if missing)
            if hasattr(self, 'compta_path_label'):
                self._compta_path_full_text = str(self._compta_base_dir)
                self.compta_path_label.configure(text=self._compta_path_full_text)
                try:
                    self._refresh_compta_path_elision()
                except Exception:
                    pass
            # Persist updated setting immediately
            try:
                self._save_settings()
            except Exception as e:
                print(f"Could not save settings: {e}")
        except Exception as e:
            print(f"Erreur lors de la mise à jour du chemin: {e}")

    def _open_compta_directory(self):
        """Open the compta base folder in the system file explorer.

        Creates the folder if it doesn't exist yet so the explorer can open it.
        """
        try:
            base = getattr(self, '_compta_base_dir', None) or self._compute_compta_base_dir()
        except Exception:
            base = Path(__file__).resolve().parent / 'dossier_compta'
        # Ensure it exists (ok to create on explicit open action)
        try:
            Path(base).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            print(f"Impossible de créer le dossier compta: {e}")
        # Open in explorer/finder/xdg
        try:
            if os.name == 'nt':
                os.startfile(str(base))
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', str(base)])
            else:
                subprocess.Popen(['xdg-open', str(base)])
        except Exception as e:
            try:
                messagebox.showerror("Ouvrir le dossier", f"Impossible d'ouvrir:\n{base}\n\n{e}")
            except Exception:
                print(f"Impossible d'ouvrir le dossier: {base} -> {e}")
    def toggle_theme(self):
        self.is_dark_mode = not self.is_dark_mode
        self.apply_theme()
        try:
            # Ensure accounts editor (if visible) is re-themed immediately
            self._refresh_accounts_editor_theme()
        except Exception:
            pass
        # Persist theme preference
        try:
            self._save_settings()
        except Exception as e:
            print(f"Could not save settings: {e}")

    # ------------------ Settings Persistence ------------------
    def _load_settings(self):
        """Load persisted settings (theme, chosen directory) from settings.json if present."""
        path = getattr(self, '_settings_path', None)
        if path is None or not path.exists():
            # No settings file => first run
            self._first_run = True
            return
        try:
            raw = path.read_text(encoding='utf-8').strip()
            if not raw:
                # Empty file => treat as first run defaults
                self._first_run = True
                return
            data = json.loads(raw)
        except Exception as e:
            print(f"Warning: could not read settings.json: {e}")
            self._first_run = True
            return
        # Theme
        is_dark = data.get("is_dark_mode")
        if isinstance(is_dark, bool):
            self.is_dark_mode = is_dark
        # Directory
        chosen_dir = data.get("chosen_parent_dir")
        if isinstance(chosen_dir, str):
            p = Path(chosen_dir)
            if p.exists() and p.is_dir():
                self._chosen_parent_dir = p
        # Accounts lists
        cur_list = data.get("current_accounts")
        if isinstance(cur_list, list):
            self.current_accounts = [str(x) for x in cur_list if isinstance(x, (str, int, float))]
        sav_list = data.get("savings_accounts")
        if isinstance(sav_list, list):
            self.savings_accounts = [str(x) for x in sav_list if isinstance(x, (str, int, float))]
            # Remove legacy default 'Épargne Principale' if it exists
            self.savings_accounts = [acc for acc in self.savings_accounts if acc.strip().lower() != "épargne principale".lower()]
        # Selected accounts
        sel_cur = data.get("selected_current_account")
        if isinstance(sel_cur, str) and sel_cur in self.current_accounts:
            self.selected_current_account = sel_cur
        sel_sav = data.get("selected_savings_account")
        if isinstance(sel_sav, str) and sel_sav in self.savings_accounts:
            try:
                self._savings_account_var.set(sel_sav)
            except Exception:
                pass
        # Provide sensible defaults if empty
        if not self.current_accounts:
            self.current_accounts = ["Compte Courant"]
        # Recompute base dir after potential directory change
        self._compta_base_dir = self._compute_compta_base_dir()
        # Persist file if any normalization applied (e.g. path missing)
        if not self._first_run:
            try:
                self._save_settings()
            except Exception:
                pass

    def _save_settings(self):
        """Persist current theme and chosen directory to settings.json."""
        path = getattr(self, '_settings_path', None)
        if path is None:
            return
        data = {
            "is_dark_mode": self.is_dark_mode,
            "chosen_parent_dir": str(self._chosen_parent_dir),
            "current_accounts": [acc for acc in self.current_accounts if str(acc).strip()],
            "savings_accounts": [acc for acc in self.savings_accounts if str(acc).strip() and str(acc).strip().lower() != "épargne principale".lower()],
            "selected_current_account": self.selected_current_account if self.selected_current_account in self.current_accounts else (self.current_accounts[0] if self.current_accounts else None),
            "selected_savings_account": (getattr(self, '_savings_account_var', StringVar()).get().strip() or None) if self.savings_accounts else None,
        }
        try:
            path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
        except Exception as e:
            print(f"Warning: could not save settings: {e}")

    def _on_close(self):
        """Handle window close event: save settings then destroy root."""
        try:
            self._save_settings()
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass

    # ------------------ Statistics Helpers ------------------
    def _update_sheet_statistics(self, ws):
        """Compute and write statistics into columns H (labels) & I (values).

        Revised logic (treat 'Épargne' as internal transfers between current & savings):
          - Non-Épargne Entrée: +current
          - Non-Épargne Sortie: -current
          - Épargne Sortie: +current, -savings  (withdraw from savings -> current)
          - Épargne Entrée: -current, +savings  (deposit into savings from current)
        This matches expectation: a 'Sortie Épargne' increases current balance and decreases savings.
        Savings balance accumulates net movements (positive => more saved, negative => withdrawn).

                Additional statistics added:
                    - Montant Total Sorties (sum of all Sortie amounts)
                    - Montant Total Entrées (sum of all Entrée amounts)
                    - Nombre Prélèvements (count where Prélèvement == 'Oui')
                    - Montant Total Prélèvements (sum of amounts where Prélèvement == 'Oui')
                """
        # Build dynamic per-account balance structures.
        # We treat column 7 ('Compte') values to infer transfers and attribution.
        # Formats written earlier:
        #   Normal (non-Épargne): '<CurrentAccountName>'
        #   Épargne Entrée:  '<Current> -> <Savings>'  (money moves current -> savings)
        #   Épargne Sortie:  '<Savings> -> <Current>'  (money moves savings -> current)
        # We do NOT require adding extra columns; we parse these directional strings.
        # Initialize dictionaries for balances; default 0.0.
        current_balances = {acc: 0.0 for acc in self.current_accounts}
        savings_balances = {acc: 0.0 for acc in self.savings_accounts}

        # Aggregate global balances (legacy single figures) as totals of dicts after loop.
        # (We will compute totals after processing all rows.)
        count_income = 0
        count_outcome = 0
        total_sorties_amount = 0.0
        total_entrees_amount = 0.0
        count_prelevements = 0
        total_prelevements_amount = 0.0

        # Per-category aggregation
        # Structure: cat_stats[category] = {"entrees_amount": float, "sorties_amount": float}
        cat_stats = {}
        # Maintain insertion order of first appearance per type (scoped inside method)
        outcome_order = []  # categories with non-zero sorties in first-seen order
        income_order = []   # categories with non-zero entrees in first-seen order

        for r in range(2, ws.max_row + 1):
            type_cell = ws.cell(row=r, column=5).value
            cat_cell = ws.cell(row=r, column=4).value
            montant_cell = ws.cell(row=r, column=3).value
            prelevement_cell = ws.cell(row=r, column=6).value
            transaction_cell = ws.cell(row=r, column=7).value
            try:
                amount = float(str(montant_cell).replace(',', '.')) if montant_cell not in (None, "") else 0.0
            except Exception:
                amount = 0.0

            # Normalize category for accent/case-insensitive savings detection
            norm_cat = ""
            if cat_cell:
                norm_cat = str(cat_cell).strip().lower()
                # Replace common accented e variants for robustness (é, è, ê)
                norm_cat = norm_cat.replace('é', 'e').replace('è', 'e').replace('ê', 'e')
            is_epargne = (norm_cat == 'epargne')

            if type_cell == "Entrée":
                count_income += 1
                total_entrees_amount += amount
                if is_epargne:
                    # New expected format for Entrée: 'Savings -> Current' (withdraw savings, add to current)
                    if isinstance(transaction_cell, str) and '->' in transaction_cell:
                        left, right = [p.strip() for p in transaction_cell.split('->', 1)]
                        sav_name, cur_name = left, right
                        if sav_name in savings_balances:
                            savings_balances[sav_name] -= amount
                        if cur_name in current_balances:
                            current_balances[cur_name] += amount
                    else:
                        if self.savings_accounts:
                            savings_balances[self.savings_accounts[0]] -= amount
                        if self.current_accounts:
                            current_balances[self.current_accounts[0]] += amount
                else:
                    # Non-savings entry increases origin account (transaction_cell may just be account name)
                    acc_name = None
                    if isinstance(transaction_cell, str):
                        name = transaction_cell.strip()
                        if name in current_balances:
                            acc_name = name
                    if acc_name is None and self.current_accounts:
                        acc_name = self.current_accounts[0]
                    if acc_name:
                        current_balances[acc_name] += amount
            elif type_cell == "Sortie":
                count_outcome += 1
                total_sorties_amount += amount
                if is_epargne:
                    # New expected format for Sortie: 'Current -> Savings' (withdraw current, add to savings)
                    if isinstance(transaction_cell, str) and '->' in transaction_cell:
                        left, right = [p.strip() for p in transaction_cell.split('->', 1)]
                        cur_name, sav_name = left, right
                        if cur_name in current_balances:
                            current_balances[cur_name] -= amount
                        if sav_name in savings_balances:
                            savings_balances[sav_name] += amount
                    else:
                        if self.current_accounts:
                            current_balances[self.current_accounts[0]] -= amount
                        if self.savings_accounts:
                            savings_balances[self.savings_accounts[0]] += amount
                else:
                    acc_name = None
                    if isinstance(transaction_cell, str):
                        name = transaction_cell.strip()
                        if name in current_balances:
                            acc_name = name
                    if acc_name is None and self.current_accounts:
                        acc_name = self.current_accounts[0]
                    if acc_name:
                        current_balances[acc_name] -= amount

            if prelevement_cell == "Oui":
                count_prelevements += 1
                total_prelevements_amount += amount

            # Track per category (ignore empty category cells)
            if cat_cell:
                if cat_cell not in cat_stats:
                    cat_stats[cat_cell] = {"entrees_amount": 0.0, "sorties_amount": 0.0}
                if type_cell == "Entrée":
                    prev = cat_stats[cat_cell]["entrees_amount"]
                    cat_stats[cat_cell]["entrees_amount"] = prev + amount
                    if prev == 0 and amount != 0 and cat_cell not in income_order:
                        income_order.append(cat_cell)
                elif type_cell == "Sortie":
                    prev = cat_stats[cat_cell]["sorties_amount"]
                    cat_stats[cat_cell]["sorties_amount"] = prev + amount
                    if prev == 0 and amount != 0 and cat_cell not in outcome_order:
                        outcome_order.append(cat_cell)

        # Compute totals from per-account dictionaries
        total_current_balance = sum(current_balances.values()) if current_balances else 0.0
        total_savings_balance = sum(savings_balances.values()) if savings_balances else 0.0

        # Build stats list in requested order:
        # 1..n: Solde Compte Courant i
        # 1..m: Solde Épargne i
        # Totals after the per-account lines
        stats = []
        # Current accounts
        for acc in self.current_accounts:
            stats.append((f"Solde {acc}", current_balances.get(acc, 0.0)))
        # Savings accounts
        for acc in self.savings_accounts:
            stats.append((f"Solde {acc}", savings_balances.get(acc, 0.0)))
        # Totals
        stats.append(("Montant Total Solde Courant", total_current_balance))
        stats.append(("Montant Total Solde Épargne", total_savings_balance))
        # Existing global metrics
        stats.extend([
            ("Nombre Sorties", count_outcome),
            ("Nombre Entrées", count_income),
            ("Montant Total Sorties", total_sorties_amount),
            ("Montant Total Entrées", total_entrees_amount),
            ("Nombre Prélèvements", count_prelevements),
            ("Montant Total Prélèvements", total_prelevements_amount),
        ])
        # NOTE (user request): per-category stats removed from H/I area.
        # We still compute cat_stats above to feed the doughnut chart and
        # populate temporary columns K/L, but we no longer append the
        # per-category breakdown lines into the stats list.

        # Clear previous stats area sized to new stats length (add a little buffer)
        clear_rows = max(16, len(stats) + 2)
        for r in range(1, clear_rows + 1):
            ws.cell(row=r, column=8).value = None
            ws.cell(row=r, column=9).value = None

        # Write labels & values
        header_font = Font(bold=True)
        ws.cell(row=1, column=8, value="Statistiques").font = header_font
        row = 2
        for label, val in stats:
            ws.cell(row=row, column=8, value=label)
            v_cell = ws.cell(row=row, column=9, value=val)
            if isinstance(val, (int, float)) and (label.startswith("Solde") or label.startswith("Montant")):
                try:
                    v_cell.number_format = "#,##0.00 [$€-fr-FR]"
                except Exception:
                    pass
            row += 1

        # Auto-adjust column widths (simple heuristic)
        for col in (8, 9):
            max_len = 0
            for r in range(1, row):
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                l = len(str(v))
                if l > max_len:
                    max_len = l
            ws.column_dimensions[get_column_letter(col)].width = max(14, min(40, max_len + 2))

        # Create category tables and charts for Outcome (Sorties) and Income (Entrées)
        try:
            # Clear K/L rows 1-40 to accommodate shifted income section
            for r in range(1, 41):
                ws.cell(row=r, column=11).value = None  # K
                ws.cell(row=r, column=12).value = None  # L

            # Row 1 title for outcome section
            ws.cell(row=1, column=11, value="Outcome")
            # Rows 2-13 outcome categories in first-seen order
            # Fill sequentially; stop if exceeding available rows
            outcome_rows_start = 2
            max_outcome_rows = 12  # rows 2..13 inclusive
            # Clear area first
            for i in range(max_outcome_rows):
                ws.cell(row=outcome_rows_start + i, column=11, value=None)
                ws.cell(row=outcome_rows_start + i, column=12, value=None)
            for idx, cat in enumerate(outcome_order):
                if idx >= max_outcome_rows:
                    break
                amt_sortie = cat_stats.get(cat, {}).get("sorties_amount", 0)
                if amt_sortie not in (0, 0.0, None):
                    row_idx = outcome_rows_start + idx
                    ws.cell(row=row_idx, column=11, value=cat)
                    vcell = ws.cell(row=row_idx, column=12, value=amt_sortie)
                    try:
                        vcell.number_format = "#,##0.00 [$€-fr-FR]"
                    except Exception:
                        pass

            # Blank separator row 14 (leave empty)
            # Rows 15-23 left intentionally blank after refactor
            # Row 15 title for income section (updated per new user request)
            income_title_cell = ws.cell(row=15, column=11, value="Income")
            try:
                income_title_cell.font = Font(bold=True, size=EXCEL_TITLE_FONT_SIZE)
            except Exception:
                pass
            # Rows 16-27 income categories in first-seen order
            income_rows_start = 16
            max_income_rows = 12  # 16..27 inclusive
            for i in range(max_income_rows):
                ws.cell(row=income_rows_start + i, column=11, value=None)
                ws.cell(row=income_rows_start + i, column=12, value=None)
            for idx, cat in enumerate(income_order):
                if idx >= max_income_rows:
                    break
                amt_entree = cat_stats.get(cat, {}).get("entrees_amount", 0)
                if amt_entree not in (0, 0.0, None):
                    row_idx = income_rows_start + idx
                    ws.cell(row=row_idx, column=11, value=cat)
                    vcell = ws.cell(row=row_idx, column=12, value=amt_entree)
                    try:
                        vcell.number_format = "#,##0.00 [$€-fr-FR]"
                    except Exception:
                        pass

            # -------- Outcome Chart (Sorties) --------
            labels_ref_out = Reference(ws, min_col=11, max_col=11, min_row=2, max_row=13)
            values_ref_out = Reference(ws, min_col=12, max_col=12, min_row=2, max_row=13)
            pie_out = PieChart()
            pie_out.title = "Sorties par Catégorie"
            pie_out.add_data(values_ref_out, titles_from_data=False)
            pie_out.set_categories(labels_ref_out)
            try:
                if pie_out.series and len(pie_out.series) > 0:
                    pie_out.series[0].title = ""
            except Exception:
                pass
            try:
                pie_out.holeSize = 55
            except Exception:
                pass
            # Determine count of actually visible outcome categories (non-empty cells written)
            visible_out_categories = [
                ws.cell(row=outcome_rows_start + i, column=11).value
                for i in range(max_outcome_rows)
                if ws.cell(row=outcome_rows_start + i, column=11).value not in (None, "")
            ]
            num_out = len(visible_out_categories)

            if DataLabelList is not None:
                try:
                    dll_o = DataLabelList()
                    dll_o.showPercent = True
                    dll_o.showCatName = True
                    dll_o.showVal = False
                    from openpyxl.chart.label import DataLabel
                    from openpyxl.drawing.text import RichText, Paragraph, ParagraphProperties, CharacterProperties
                    for i in range(num_out):
                        dl = DataLabel(idx=i)
                        dl.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1400)))])
                        dll_o.dLbl.append(dl)
                    pie_out.dataLabels = dll_o
                except Exception:
                    pass
            # Palette for outcome
            palette_out = [
                "FF5722","4CAF50","2196F3","9C27B0","FFC107","009688","E91E63","3F51B5","795548","607D8B","8BC34A","FF9800"
            ]
            try:
                if pie_out.series:
                    ser_o = pie_out.series[0]
                    for i in range(len(CATEGORY_VALUES)):
                        from openpyxl.chart.series import DataPoint
                        try:
                            pt = ser_o.dPt[i]
                        except IndexError:
                            pt = DataPoint(idx=i)
                            ser_o.dPt.append(pt)
                        try:
                            color_hex = palette_out[i % len(palette_out)]
                            pt.graphicalProperties.solidFill = color_hex
                        except Exception:
                            pass
            except Exception:
                pass
            try:
                if getattr(pie_out, 'legend', None) is None:
                    from openpyxl.chart.legend import Legend as _Legend
                    pie_out.legend = _Legend()
                pie_out.legend.position = 'r'
                from openpyxl.drawing.text import RichText, Paragraph, ParagraphProperties, CharacterProperties
                pie_out.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1600)))])
            except Exception:
                pass
            pie_out.height = CHART_HEIGHT
            pie_out.width = CHART_WIDTH
            try:
                from openpyxl.drawing.text import ParagraphProperties, CharacterProperties
                if getattr(pie_out, 'title', None) and getattr(pie_out.title, 'tx', None) and getattr(pie_out.title.tx, 'rich', None):
                    for p in pie_out.title.tx.rich.p:
                        if p.pPr is None:
                            p.pPr = ParagraphProperties()
                        p.pPr.defRPr = CharacterProperties(sz=2000)
            except Exception:
                pass

            # Remove previous outcome chart(s)
            try:
                to_remove = []
                for obj in ws._charts:
                    if getattr(obj, 'title', None) and 'Sorties par Catégorie' in str(obj.title):
                        to_remove.append(obj)
                for obj in to_remove:
                    ws._charts.remove(obj)
            except Exception:
                pass
            ws.add_chart(pie_out, "O2")

            # -------- Summary Bar Chart (Income vs Outcome) --------
            try:
                # Place total outcome (sorties) and total income (entrées) into temporary cells for chart source
                # M/N (13/14) previously used but conflicted with charts; now using Y/Z (25/26)
                # Clear legacy temp cells if present (columns 13/14 rows 2-3)
                for legacy_col in (13, 14):
                    for legacy_row in (2, 3):
                        try:
                            ws.cell(row=legacy_row, column=legacy_col).value = None
                        except Exception:
                            pass
                LABEL_COL = 25  # Y
                VALUE_COL = 26  # Z
                ws.cell(row=2, column=LABEL_COL, value="Outcome Total")
                ws.cell(row=3, column=LABEL_COL, value="Income Total")
                ws.cell(row=2, column=VALUE_COL, value=total_sorties_amount)
                ws.cell(row=3, column=VALUE_COL, value=total_entrees_amount)
                bar_values_ref = Reference(ws, min_col=VALUE_COL, max_col=VALUE_COL, min_row=2, max_row=3)
                bar_labels_ref = Reference(ws, min_col=LABEL_COL, max_col=LABEL_COL, min_row=2, max_row=3)
                bar_chart = BarChart()
                bar_chart.title = "Entrées / Sorties"
                bar_chart.add_data(bar_values_ref, titles_from_data=False)
                bar_chart.set_categories(bar_labels_ref)
                bar_chart.y_axis.title = "Montant (€)"
                bar_chart.x_axis.title = "Type"
                bar_chart.type = "col"
                bar_chart.style = 2
                # Adjust size similar proportion to pie charts but a bit narrower
                bar_chart.width = CHART_WIDTH * 0.75
                bar_chart.height = CHART_HEIGHT
                # Color bars: row 2 (Outcome) red, row 3 (Income) green
                try:
                    if bar_chart.series:
                        ser = bar_chart.series[0]
                        from openpyxl.chart.series import DataPoint
                        # Outcome (index 0)
                        dp0 = DataPoint(idx=0)
                        dp0.graphicalProperties.solidFill = "C62828"  # red
                        # Income (index 1)
                        dp1 = DataPoint(idx=1)
                        dp1.graphicalProperties.solidFill = "2E7D32"  # green
                        ser.dPt.append(dp0)
                        ser.dPt.append(dp1)
                except Exception:
                    pass
                # Remove previous bar chart if exists
                try:
                    to_remove = []
                    for obj in ws._charts:
                        if getattr(obj, 'title', None) and 'Entrées / Sorties' in str(obj.title):
                            to_remove.append(obj)
                    for obj in to_remove:
                        ws._charts.remove(obj)
                except Exception:
                    pass
                # Place bar chart to the right of outcome pie (which starts at O2).
                # Outcome pie roughly spans O2 -> (depending on width). We'll anchor bar chart at "Z2" for spacing.
                # Anchor the bar chart further right (AC2) so it doesn't overlap with the Y/Z data columns
                ws.add_chart(bar_chart, "AC2")
            except Exception as e:
                print(f"Warning: could not create income/outcome bar chart: {e}")

                # Continue even if bar chart fails


            # -------- Income Chart (Entrées) --------
            labels_ref_in = Reference(ws, min_col=11, max_col=11, min_row=16, max_row=27)
            values_ref_in = Reference(ws, min_col=12, max_col=12, min_row=16, max_row=27)
            pie_in = PieChart()
            pie_in.title = "Entrées par Catégorie"
            pie_in.add_data(values_ref_in, titles_from_data=False)
            pie_in.set_categories(labels_ref_in)
            try:
                if pie_in.series and len(pie_in.series) > 0:
                    pie_in.series[0].title = ""
            except Exception:
                pass
            try:
                pie_in.holeSize = 55
            except Exception:
                pass
            # Determine count of actually visible income categories (non-empty cells written)
            visible_in_categories = [
                ws.cell(row=income_rows_start + i, column=11).value
                for i in range(max_income_rows)
                if ws.cell(row=income_rows_start + i, column=11).value not in (None, "")
            ]
            num_in = len(visible_in_categories)

            if DataLabelList is not None:
                try:
                    dll_i = DataLabelList()
                    dll_i.showPercent = True
                    dll_i.showCatName = True
                    dll_i.showVal = False
                    from openpyxl.chart.label import DataLabel
                    from openpyxl.drawing.text import RichText, Paragraph, ParagraphProperties, CharacterProperties
                    for i in range(num_in):
                        dl = DataLabel(idx=i)
                        dl.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1400)))])
                        dll_i.dLbl.append(dl)
                    pie_in.dataLabels = dll_i
                except Exception:
                    pass
            palette_in = [
                "2196F3","4CAF50","FFC107","9C27B0","FF5722","009688","E91E63","3F51B5","795548","607D8B","8BC34A","FF9800"
            ]
            try:
                if pie_in.series:
                    ser_i = pie_in.series[0]
                    from openpyxl.chart.series import DataPoint
                    for i in range(num_in):
                        try:
                            pt = ser_i.dPt[i]
                        except IndexError:
                            pt = DataPoint(idx=i)
                            ser_i.dPt.append(pt)
                        try:
                            color_hex = palette_in[i % len(palette_in)]
                            pt.graphicalProperties.solidFill = color_hex
                        except Exception:
                            pass
            except Exception:
                pass
            try:
                if getattr(pie_in, 'legend', None) is None:
                    from openpyxl.chart.legend import Legend as _Legend
                    pie_in.legend = _Legend()
                pie_in.legend.position = 'r'
                from openpyxl.drawing.text import RichText, Paragraph, ParagraphProperties, CharacterProperties
                pie_in.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1600)))])
            except Exception:
                pass
            pie_in.height = CHART_HEIGHT
            pie_in.width = CHART_WIDTH
            try:
                from openpyxl.drawing.text import ParagraphProperties, CharacterProperties
                if getattr(pie_in, 'title', None) and getattr(pie_in.title, 'tx', None) and getattr(pie_in.title.tx, 'rich', None):
                    for p in pie_in.title.tx.rich.p:
                        if p.pPr is None:
                            p.pPr = ParagraphProperties()
                        p.pPr.defRPr = CharacterProperties(sz=2000)
            except Exception:
                pass
            # Remove previous income chart(s)
            try:
                to_remove = []
                for obj in ws._charts:
                    if getattr(obj, 'title', None) and 'Entrées par Catégorie' in str(obj.title):
                        to_remove.append(obj)
                for obj in to_remove:
                    ws._charts.remove(obj)
            except Exception:
                pass
            ws.add_chart(pie_in, "O15")  # repositioned to align with new income title row (row 15)
            # Ensure uniform row heights so blank spacer rows (14-23) aren't visually compressed
            try:
                # Choose a height that fits 18-20pt fonts comfortably
                uniform_height = 28  # points
                for r in range(1, 41):  # cover both outcome & income table zones
                    rd = ws.row_dimensions[r]
                    rd.height = uniform_height
            except Exception:
                pass
        except Exception as e:
            print(f"Warning: could not create category charts: {e}")

    def _apply_sheet_fonts(self, ws):
        """Set header row fonts (bold size EXCEL_TITLE_FONT_SIZE) and remaining populated rows
        font size EXCEL_BODY_FONT_SIZE. Preserves existing font colors (already applied for semantic
        coloring) by reusing the color attribute.
        """
        try:
            max_col = ws.max_column
            max_row = ws.max_row
            # Header row (row 1)
            for c in range(1, max_col + 1):
                cell = ws.cell(row=1, column=c)
                if cell.value in (None, ""):
                    continue
                prev_font = cell.font
                cell.font = Font(bold=True, size=EXCEL_TITLE_FONT_SIZE, color=getattr(prev_font, 'color', None))
                try:
                    cell.alignment = Alignment(horizontal='right')
                except Exception:
                    pass

            # Body rows
            for r in range(2, max_row + 1):
                # Skip entirely empty rows quickly
                if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, max_col + 1)):
                    continue
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value in (None, ""):
                        continue
                    # Preserve Income title at K21 (row 21, col 11)
                    if r == 15 and c == 11 and cell.value == "Income":
                        prev_color = getattr(cell.font, 'color', None)
                        cell.font = Font(bold=True, size=EXCEL_TITLE_FONT_SIZE, color=prev_color)
                        continue
                    prev_font = cell.font
                    cell.font = Font(bold=False, size=EXCEL_BODY_FONT_SIZE, color=getattr(prev_font, 'color', None))
                    try:
                        cell.alignment = Alignment(horizontal='right')
                    except Exception:
                        pass
            # Re-assert Income title styling explicitly (safety)
            try:
                income_cell = ws.cell(row=15, column=11)
                if income_cell.value == "Income":
                    prev_color = getattr(income_cell.font, 'color', None)
                    income_cell.font = Font(bold=True, size=EXCEL_TITLE_FONT_SIZE, color=prev_color)
                    try:
                        income_cell.alignment = Alignment(horizontal='right')
                    except Exception:
                        pass
            except Exception:
                pass
        except Exception as e:
            print(f"Font sizing error: {e}")

    def _find_next_transaction_row(self, ws):
        """Return the next row index for a new transaction.

        Ignores statistics in columns H/I so they don't create artificial gaps.
        Scans only columns 1..6 (A-F) for existing data rows.
        """
        last_data_row = 1  # header at row 1
        max_row = ws.max_row
        for r in range(2, max_row + 1):
            if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, 7)):
                last_data_row = r
        return last_data_row + 1

    def _auto_adjust_transaction_columns(self, ws):
        """Auto size columns A-F based on content length with more generous widths.

        New adaptive rules (character-based Excel width units):
          A Date:           min 11, max 16 (dates fixed length but give breathing room)
          B Libellé:        min 20, soft target = max_len+2, hard max 80 (long labels)
          C Montant:        min 12, max 20 (allow larger numbers / decimals)
          D Catégorie:      min 16, max 32 (French words & accents)
          E Type:           min 10, max 16 ("Entrée" / "Sortie")
          F Prélèvement:    min 14, max 22 ("Oui" / "Non" plus header)

        Strategy:
          - Scan only data rows (A-F) ignoring stats area.
          - Compute max text length per column.
          - Add padding ( +2 or +3 for Libellé ).
          - Clamp inside new bounds.
          - If Libellé contains very long outliers ( > 70 chars ), cap at 80 but still wide.
        """
        config = {
            1: {"min": 11, "max": 16, "pad": 2},   # Date
            2: {"min": 20, "max": 80, "pad": 3},   # Libellé
            3: {"min": 12, "max": 20, "pad": 2},   # Montant
            4: {"min": 16, "max": 32, "pad": 2},   # Catégorie
            5: {"min": 10, "max": 16, "pad": 2},   # Type
            6: {"min": 14, "max": 22, "pad": 2},   # Prélèvement
        }

        last_data_row = self._find_next_transaction_row(ws) - 1
        if last_data_row < 1:
            return

        for col in range(1, 7):
            cfg = config[col]
            max_len = 0
            # Include header row for measuring
            for r in range(1, last_data_row + 1):
                val = ws.cell(row=r, column=col).value
                if val is None:
                    continue
                s = str(val)
                # For Montant, remove thousand separators just for width logic
                if col == 3:
                    s = s.replace('€', '').strip()
                length = len(s)
                if length > max_len:
                    max_len = length

            if max_len == 0:
                target = cfg["min"]
            else:
                target = max_len + cfg["pad"]

            # Special handling for Libellé: don't exceed hard max but allow wide content
            if col == 2 and max_len > 70:
                target = cfg["max"]

            # Clamp
            if target < cfg["min"]:
                target = cfg["min"]
            if target > cfg["max"]:
                target = cfg["max"]

            ws.column_dimensions[get_column_letter(col)].width = target
    def _apply_fixed_widths(self, ws):
        """Set fixed widths: A-F => 20, H=40, I=20, K=35 (chart/data area).

        Leaves other columns unchanged if present. Safe to call multiple times.
        """
        try:
            for col in range(1, 7):  # A-F
                ws.column_dimensions[get_column_letter(col)].width = 20
            # Column G (Transaction) wider for directional strings
            ws.column_dimensions[get_column_letter(7)].width = 40
            # Stats columns H (8) & I (9)
            ws.column_dimensions[get_column_letter(8)].width = 40  # H
            ws.column_dimensions[get_column_letter(9)].width = 20  # I
            # Category chart source data column K (11)
            ws.column_dimensions[get_column_letter(11)].width = 30  # K
            ws.column_dimensions[get_column_letter(12)].width = 20  # L
            ws.column_dimensions[get_column_letter(13)].width = 12  # M
            ws.column_dimensions[get_column_letter(25)].width = 20  # Y
            ws.column_dimensions[get_column_letter(26)].width = 20  # Z
        except Exception as e:
            print(f"Fixed width error: {e}")

    def _normalize_date_column(self, ws):
        """Convert textual DD-MM-YYYY (or DD/MM/YYYY) values in column A to real date objects.

        Safe to run repeatedly; skips cells already recognized as dates.
        """
        pattern = re.compile(r"^(\d{2})[-/](\d{2})[-/](\d{4})$")
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=1)
            val = cell.value
            if val is None:
                continue
            # If already a date/datetime, ensure number format and continue
            if isinstance(val, (datetime, date)):
                try:
                    cell.number_format = "dd/mm/yyyy"
                except Exception:
                    pass
                continue
            if isinstance(val, str):
                m = pattern.match(val.strip())
                if not m:
                    continue
                try:
                    # Accept either '-' or '/' as separator in stored text
                    norm = val.strip().replace('/', '-')
                    d = datetime.strptime(norm, "%d-%m-%Y").date()
                    cell.value = d
                    cell.number_format = "dd/mm/yyyy"
                except Exception:
                    continue

    def _on_date_entry_click(self):
        """Handle click on date entry - prevent immediate calendar closing"""
        self._calendar_opening = True
        self._open_calendar()
        # Reset flag after a short delay
        self.root.after(100, lambda: setattr(self, '_calendar_opening', False))

    def _open_calendar(self):
        if self.calendar_win is not None and self.calendar_win.winfo_exists():
            self.calendar_win.lift()
            return
        self.calendar_win = Toplevel(self.root)

        # Windows-compatible setup
        self.calendar_win.title("Select Date")
        self.calendar_win.resizable(False, False)
        self.calendar_win.transient(self.root)

        # Try to set topmost, but don't fail if it doesn't work on Windows
        try:
            self.calendar_win.attributes("-topmost", True)
        except Exception:
            pass

        # Force window updates to ensure accurate positioning
        self.root.update_idletasks()
        self.calendar_win.update_idletasks()

        # Small delay to ensure widgets are properly rendered
        self.root.after(1, self._do_position_calendar)

    def _do_position_calendar(self):
        if self.calendar_win is None or not self.calendar_win.winfo_exists():
            return

        # Get entry field position
        x = self.date_entry.winfo_rootx()
        y = self.date_entry.winfo_rooty() + self.date_entry.winfo_height()

        # Adjust for Windows DPI and screen boundaries
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Calendar dimensions
        cal_width = 260
        cal_height = 250

        # Ensure calendar stays on screen
        if x + cal_width > screen_width:
            x = screen_width - cal_width - 10
        if y + cal_height > screen_height:
            y = self.date_entry.winfo_rooty() - cal_height

        self.calendar_win.geometry(f"{cal_width}x{cal_height}+{x}+{y}")
        # Populate the calendar (header + body) – previously missing, causing empty popup
        try:
            # Sync displayed month/year to current entry value if parsable
            try:
                self._sync_calendar_to_entry()
            except Exception:
                pass
            # (Re)build header and body every open to reflect any changes
            try:
                self._build_calendar_header()
            except Exception:
                pass
            try:
                self._build_calendar_body()
            except Exception:
                pass
            # Apply theme to freshly created widgets
            try:
                self._refresh_calendar_theme()
            except Exception:
                pass
            # Ensure focus-out listener is attached so clicking elsewhere closes popup
            try:
                self.calendar_win.bind("<FocusOut>", self._on_calendar_focus_out)
            except Exception:
                pass
        except Exception:
            # Soft-fail; avoid crashing the main app if calendar build has an issue
            pass

    def _build_calendar_header(self):
        for child in self.calendar_win.winfo_children():
            child.destroy()

        theme = self.current_theme
        header = Frame(self.calendar_win, bg=theme["bg"])
        header.pack(fill="x", pady=4)

        prev_btn = Button(
            header,
            text="◀",
            width=2,
            command=self._prev_month,
            relief="flat",
            borderwidth=0,
            bg=theme["entry_bg"],
            fg=theme["entry_fg"],
            activebackground=theme["toggle_track"],
            activeforeground=theme["toggle_thumb"]
        )
        prev_btn.pack(side="left", padx=4)

        next_btn = Button(
            header,
            text="▶",
            width=2,
            command=self._next_month,
            relief="flat",
            borderwidth=0,
            bg=theme["entry_bg"],
            fg=theme["entry_fg"],
            activebackground=theme["toggle_track"],
            activeforeground=theme["toggle_thumb"]
        )
        next_btn.pack(side="right", padx=4)

        month_label = Label(
            header,
            text=f"{_calendar.month_name[self._calendar_month]} {self._calendar_year}",
            anchor="center",
            bg=theme["bg"],
            fg=theme["fg"]
        )
        month_label.pack(fill="x")
        self._calendar_header_widgets = (header, prev_btn, next_btn, month_label)

    def _build_calendar_body(self):
        existing = getattr(self, "_calendar_day_frame", None)
        if existing and existing.winfo_exists():
            existing.destroy()

        theme = self.current_theme
        day_frame = Frame(self.calendar_win, bg=theme["bg"])
        day_frame.pack(fill="both", expand=True, padx=6, pady=(0, 6))
        self._calendar_day_frame = day_frame

        # Day headers
        for idx, wd in enumerate(["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]):
            lbl = Label(
                day_frame,
                text=wd,
                anchor="center",
                bg=theme["bg"],
                fg=theme["fg"]
            )
            lbl.grid(row=0, column=idx, padx=2, pady=2, sticky="nsew")

        for c in range(7):
            day_frame.grid_columnconfigure(c, weight=1, uniform="day")

        first_wd, days_in_month = _calendar.monthrange(self._calendar_year, self._calendar_month)
        row = 1
        col = first_wd

        for d in range(1, days_in_month + 1):
            current_date = date(self._calendar_year, self._calendar_month, d)
            ds_display = current_date.strftime("%d-%m-%Y")
            is_selected = self._selected_date == current_date

            if is_selected:
                btn = Button(
                    day_frame,
                    text=str(d),
                    width=2,
                    relief="flat",
                    borderwidth=0,
                    command=lambda ds=ds_display: self._select_date(ds),
                    bg=theme["toggle_track_active"],
                    fg=theme["toggle_thumb"],
                    activebackground=theme["toggle_track_active"],
                    activeforeground=theme["toggle_thumb"]
                )
                btn._is_selected = True
            else:
                btn = Button(
                    day_frame,
                    text=str(d),
                    width=2,
                    relief="flat",
                    borderwidth=0,
                    command=lambda ds=ds_display: self._select_date(ds),
                    bg=theme["entry_bg"],
                    fg=theme["entry_fg"],
                    activebackground=theme["toggle_track"],
                    activeforeground=theme["toggle_thumb"]
                )

            btn.grid(row=row, column=col, padx=1, pady=1, sticky="nsew")
            col += 1
            if col > 6:
                col = 0
                row += 1

        for r in range(row + 1):
            day_frame.grid_rowconfigure(r, weight=1)

    def _select_date(self, ds: str):
        self.date_var.set(ds)
        try:
            self._selected_date = datetime.strptime(ds, "%d-%m-%Y").date()
        except Exception:
            self._selected_date = None
        if self.calendar_win and self.calendar_win.winfo_exists():
            self._build_calendar_body()
            self._refresh_calendar_theme()
            # Recolor accounts button
            if hasattr(self, 'accounts_button'):
                try:
                    self.accounts_button.configure(bg=self.current_theme["bg"], fg=self.current_theme["fg"], activebackground=self.current_theme["bg"], activeforeground=self.current_theme["fg"])
                except Exception:
                    pass
        self._close_calendar()

    def _sync_calendar_to_entry(self):
        value = self.date_var.get().strip()
        try:
            dt = datetime.strptime(value, "%d-%m-%Y")
            self._calendar_year = dt.year
            self._calendar_month = dt.month
            self._selected_date = dt.date()
        except Exception:
            today = date.today()
            self._calendar_year = today.year
            self._calendar_month = today.month
            self._selected_date = None

    # --- Calendar navigation & theming (restored) ---------------------------------
    def _prev_month(self):
        """Go to previous month in the calendar widget and rebuild UI."""
        self._calendar_month -= 1
        if self._calendar_month < 1:
            self._calendar_month = 12
            self._calendar_year -= 1
        if self.calendar_win and self.calendar_win.winfo_exists():
            self._build_calendar_header()
            self._build_calendar_body()
            self._refresh_calendar_theme()

    def _next_month(self):
        """Go to next month in the calendar widget and rebuild UI."""
        self._calendar_month += 1
        if self._calendar_month > 12:
            self._calendar_month = 1
            self._calendar_year += 1
        if self.calendar_win and self.calendar_win.winfo_exists():
            self._build_calendar_header()
            self._build_calendar_body()
            self._refresh_calendar_theme()

    def _refresh_calendar_theme(self):
        """Apply the current theme colors to the calendar popup (if open)."""
        if not (self.calendar_win and self.calendar_win.winfo_exists()):
            return
        theme = self.current_theme
        try:
            self.calendar_win.configure(bg=theme["bg"])  # Window background
        except Exception:
            pass
        # Header widgets: stored as (header_frame, prev_btn, next_btn, month_label)
        header_widgets = getattr(self, "_calendar_header_widgets", None)
        if header_widgets:
            header, prev_btn, next_btn, month_label = header_widgets
            for w in (header,):
                try:
                    w.configure(bg=theme["bg"])
                except Exception:
                    pass
            for btn in (prev_btn, next_btn):
                try:
                    btn.configure(
                        bg=theme["entry_bg"],
                        fg=theme["entry_fg"],
                        activebackground=theme["toggle_track"],
                        activeforeground=theme["toggle_thumb"]
                    )
                except Exception:
                    pass
            try:
                month_label.configure(bg=theme["bg"], fg=theme["fg"])
            except Exception:
                pass
        # Day buttons inside self._calendar_day_frame
        day_frame = getattr(self, "_calendar_day_frame", None)
        if day_frame and day_frame.winfo_exists():
            for child in day_frame.winfo_children():
                # Skip the weekday header labels (they have no _is_selected attr)
                if isinstance(child, Button):
                    is_sel = getattr(child, "_is_selected", False)
                    try:
                        if is_sel:
                            child.configure(
                                bg=theme["toggle_track_active"],
                                fg=theme["toggle_thumb"],
                                activebackground=theme["toggle_track_active"],
                                activeforeground=theme["toggle_thumb"]
                            )
                        else:
                            child.configure(
                                bg=theme["entry_bg"],
                                fg=theme["entry_fg"],
                                activebackground=theme["toggle_track"],
                                activeforeground=theme["toggle_thumb"]
                            )
                    except Exception:
                        pass

    # ------------------ Accounts Management (class-level methods) ------------------
    def _open_accounts_manager(self):
        if hasattr(self, '_accounts_win') and self._accounts_win and self._accounts_win.winfo_exists():
            try:
                self._accounts_win.lift()
                return
            except Exception:
                pass
        win = Toplevel(self.root)
        self._accounts_win = win
        win.title("Gestion des Comptes")
        try:
            win.configure(bg=self.current_theme["bg"])
        except Exception:
            pass
        win.geometry("420x340")
        frm_cur = Frame(win, bg=self.current_theme["bg"])
        frm_sav = Frame(win, bg=self.current_theme["bg"])
        frm_buttons = Frame(win, bg=self.current_theme["bg"])
        frm_cur.pack(fill="both", expand=True, padx=12, pady=(12,6))
        frm_sav.pack(fill="both", expand=True, padx=12, pady=(0,6))
        frm_buttons.pack(fill="x", padx=12, pady=8)
        lbl_cur = Label(frm_cur, text="Comptes Courants", bg=self.current_theme["bg"], fg=self.current_theme["fg"], font=(FONT_FAMILY, 12, 'bold'))
        lbl_cur.pack(anchor="w")
        self._cur_listbox = Listbox(frm_cur, height=5, activestyle="none")
        self._cur_listbox.pack(fill="x", pady=4)
        for name in self.current_accounts:
            self._cur_listbox.insert('end', name)
        cur_add_frame = Frame(frm_cur, bg=self.current_theme["bg"])
        cur_add_frame.pack(fill="x", pady=(2,4))
        self._cur_new_var = StringVar()
        cur_entry = Entry(cur_add_frame, textvariable=self._cur_new_var, width=24)
        cur_entry.pack(side="left", padx=(0,6))
        cur_btn = Button(cur_add_frame, text="Ajouter", command=self._add_current_account, relief="flat", bg=self.current_theme["toggle_track_active"], fg=self.current_theme["toggle_thumb"], cursor="hand2")
        cur_btn.pack(side="left")
        lbl_sav = Label(frm_sav, text="Comptes Épargne", bg=self.current_theme["bg"], fg=self.current_theme["fg"], font=(FONT_FAMILY, 12, 'bold'))
        lbl_sav.pack(anchor="w")
        self._sav_listbox = Listbox(frm_sav, height=5, activestyle="none")
        self._sav_listbox.pack(fill="x", pady=4)
        for name in self.savings_accounts:
            self._sav_listbox.insert('end', name)
        sav_add_frame = Frame(frm_sav, bg=self.current_theme["bg"])
        sav_add_frame.pack(fill="x", pady=(2,4))
        self._sav_new_var = StringVar()
        sav_entry = Entry(sav_add_frame, textvariable=self._sav_new_var, width=24)
        sav_entry.pack(side="left", padx=(0,6))
        sav_btn = Button(sav_add_frame, text="Ajouter", command=self._add_savings_account, relief="flat", bg=self.current_theme["toggle_track_active"], fg=self.current_theme["toggle_thumb"], cursor="hand2")
        sav_btn.pack(side="left")
        close_btn = Button(frm_buttons, text="Fermer", command=win.destroy, relief="flat", bg=self.current_theme["entry_bg"], fg=self.current_theme["entry_fg"], cursor="hand2")
        close_btn.pack(side="right")

    def _add_current_account(self):
        name = (getattr(self, '_cur_new_var', StringVar()).get() or "").strip()
        if not name:
            return
        if name in self.current_accounts:
            return
        self.current_accounts.append(name)
        try:
            self._cur_listbox.insert('end', name)
        except Exception:
            pass
        try:
            self._cur_new_var.set("")
        except Exception:
            pass
        try:
            self._save_settings()
        except Exception:
            pass

    def _add_savings_account(self):
        name = (getattr(self, '_sav_new_var', StringVar()).get() or "").strip()
        if not name:
            return
        if name in self.savings_accounts:
            return
        self.savings_accounts.append(name)
        try:
            self._sav_listbox.insert('end', name)
        except Exception:
            pass
        try:
            self._sav_new_var.set("")
        except Exception:
            pass
        try:
            self._save_settings()
        except Exception:
            pass

    def _close_calendar(self):
        if self.calendar_win is not None:
            try:
                self.calendar_win.destroy()
            except Exception:
                pass
            self.calendar_win = None

    def _on_calendar_focus_out(self, event):
        self.root.after(100, self._check_calendar_focus)

    def _check_calendar_focus(self):
        if self.calendar_win is None or not self.calendar_win.winfo_exists():
            return
        try:
            focused = self.root.focus_get()
            if focused is None or str(focused).find(str(self.calendar_win)) == -1:
                if not str(focused).find(str(self.date_entry)) >= 0:
                    self._close_calendar()
        except Exception:
            pass

    def _maybe_close_calendar(self, event):
        if self.calendar_win is None or not self.calendar_win.winfo_exists():
            return

        if getattr(self, '_calendar_opening', False):
            return

        widget = event.widget

        if widget is self.date_entry:
            return

        try:
            current_widget = widget
            while current_widget:
                if current_widget == self.calendar_win:
                    return
                try:
                    current_widget = current_widget.master
                except AttributeError:
                    break

            widget_path = str(widget)
            calendar_path = str(self.calendar_win)
            if widget_path.startswith(calendar_path):
                return

        except Exception:
            pass

        self._close_calendar()

    # ------------------ Account Selectors (added) ------------------
    def _inject_account_selectors(self):
        """Create or refresh account selection comboboxes.

        Current account selector lives in header.
        Savings selector shows only when category == 'Épargne'.
        """
        try:
            # Current account combobox
            if not hasattr(self, '_current_account_cb'):
                self._current_account_cb = ttk.Combobox(
                    self.header_frame,
                    textvariable=self._current_account_var,
                    values=self.current_accounts,
                    state='readonly',
                    height=4,  # fewer visible rows in dropdown
                    width=14,  # narrower field
                    style=self.combobox_style,  # unified style
                )
                # Slightly reduced vertical padding to keep header compact
                self._current_account_cb.pack(side='left', padx=(12,8), pady=2)
                self._current_account_cb.bind('<<ComboboxSelected>>', self._on_current_account_change)
                try:
                    self._bind_combobox_popdown_theming(self._current_account_cb)
                except Exception:
                    pass
            else:
                # Ensure the dropdown height stays compact and values are refreshed
                self._current_account_cb.configure(values=self.current_accounts, height=4, width=14, style=self.combobox_style)
            if not self._current_account_var.get() and self.current_accounts:
                self._current_account_var.set(self.current_accounts[0])
                self.selected_current_account = self.current_accounts[0]
            # Savings combobox (in entry frame) hidden unless Épargne
            if not hasattr(self, '_savings_account_cb'):
                self._savings_account_cb = ttk.Combobox(
                    self.entry_frame,
                    textvariable=self._savings_account_var,
                    values=self.savings_accounts,
                    state='readonly',
                    width=20,
                    style=self.combobox_style,
                )
                self._savings_account_cb.grid(row=2, column=1, padx=10, pady=(2,0), sticky='new')
                try:
                    self._bind_combobox_popdown_theming(self._savings_account_cb)
                except Exception:
                    pass
            # Bind category selection to visibility evaluation
            try:
                self.category_combobox.bind('<<ComboboxSelected>>', self._evaluate_savings_selector_visibility)
            except Exception:
                pass
            self._evaluate_savings_selector_visibility()
        except Exception as e:
            print(f"Account selector init error: {e}")

    def _evaluate_savings_selector_visibility(self, *_):
        try:
            if self.category_var.get().strip() == 'Épargne':
                self._savings_account_cb.configure(values=self.savings_accounts)
                self._savings_account_cb.grid()
            else:
                self._savings_account_var.set("")
                self._savings_account_cb.grid_remove()
        except Exception:
            pass

    def _on_current_account_change(self, *_):
        sel = self._current_account_var.get().strip()
        if sel and sel in self.current_accounts:
            self.selected_current_account = sel
            try:
                self._save_settings()
            except Exception:
                pass

if __name__ == "__main__":
    root = Tk()
    app = App(root)

    root.bind("<Escape>", lambda e: root.quit())
    root.mainloop()
